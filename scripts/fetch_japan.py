#!/usr/bin/env python3
"""
Fetch Japan vehicle registration data from JADA and update data/Japan.csv.

Usage
-----
    python scripts/fetch_japan.py [--year YEAR] [--month MONTH] \
        [--xlsx-url URL_OR_PATH] [--pdf-url URL_OR_PATH] \
        [--csv PATH] [--force]

* --year / --month  Override the target month (default: previous calendar month).
* --xlsx-url        Direct URL/path to the JADA monthly XLSX (preferred parser).
* --pdf-url         Direct URL/path to the JADA monthly PDF (fallback parser).
* --csv             Target CSV (default: data/Japan.csv).
* --force           Re-process even if the target period already exists.

Invoked by .github/workflows/fetch-japan.yml on a daily cron from the 1st of
each month onward, plus manual workflow_dispatch. When the CSV changes, the
workflow commits data/Japan.csv and triggers render-country.yml for Japan.

Data source
-----------
JADA (日本自動車販売協会連合会 — Japan Automobile Dealers Association) publishes
on https://www.jada.or.jp/pages/342/ a monthly "燃料別メーカー別登録台数（乗用車）"
file (Fuel-type, Manufacturer-wise Registrations — passenger cars). Each
publication is a rolling 4-month rollup: one sheet/page per month, current
month at the top. We extract the most recent (or a requested) month.

Both formats (PDF and XLSX) carry the same data. XLSX is preferred because
the cell layout is machine-readable; PDF is a fallback for cases where only
the PDF was published.

Vehicle scope
-------------
登録車 (registered passenger cars) only. The footer note "２．軽自動車は含み
ません" makes this explicit — kei cars are NOT included. The historical rows
in data/Japan.csv match this scope (totals ~180-280k/month).

CSV layout
----------
Existing data/Japan.csv columns:
    period,time_interval,variant,source,BEV,PHEV,HEV,PETROL,DIESEL,OTHERS,TOTAL,notes

JADA column                           → CSV column
    ガソリン (Gasoline)               → PETROL
    ＨＶ (Hybrid)                      → HEV
    ＰＨＶ (Plug-in Hybrid)            → PHEV
    ディーゼル (Diesel)                → DIESEL
    ＥＶ (Electric)                    → BEV
    ＦＣＶ (Fuel-cell) + その他(*)     → OTHERS  (matches historical CSV: e.g.
                                                  2026-01 OTHERS=80 = FCV 79 + その他 1)
    合計 (Total)                       → TOTAL

Sanity check: PETROL+HEV+PHEV+DIESEL+BEV+OTHERS must equal TOTAL.

Per the project rule we only ever write the most recent month; older rows
are never touched, even if a later JADA file would adjust them.

Auto-discovery
--------------
JADA's hosting blocks some cloud IP ranges with HTTP 403
"x-deny-reason: host_not_allowed". If the index page or the file URL
cannot be fetched from CI, pass --xlsx-url / --pdf-url manually (or via
workflow_dispatch inputs). Local file paths are also accepted.
"""
import argparse
import csv
import io
import re
import sys
from datetime import date
from pathlib import Path

import requests

JADA_PAGE = "https://www.jada.or.jp/pages/342/"
JADA_HOST = "https://www.jada.or.jp"

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja-JP,ja;q=0.9,en;q=0.8",
    "Referer": "https://www.jada.or.jp/",
}

CSV_COLUMNS = [
    "period", "time_interval", "variant", "source",
    "BEV", "PHEV", "HEV", "PETROL", "DIESEL", "OTHERS",
    "TOTAL", "notes",
]

# Row label that flags the all-makes total in each monthly sheet/page.
TOTAL_ROW_LABEL = "乗用車計"

# Sheet/page title pattern: "燃料別メーカー別登録台数（乗用車）　2026年4月"
TITLE_RE = re.compile(r"燃料別メーカー別登録台数.*?(\d{4})年\s*(\d{1,2})月")


def previous_month(today: date) -> tuple[int, int]:
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


def latest_period(csv_path: str) -> str | None:
    if not Path(csv_path).exists():
        return None
    with open(csv_path, newline="", encoding="utf-8") as f:
        periods = [row["period"] for row in csv.DictReader(f)]
    return max(periods) if periods else None


def load_bytes(url_or_path: str) -> bytes:
    if url_or_path.startswith("http://") or url_or_path.startswith("https://"):
        print(f"Downloading: {url_or_path}")
        resp = requests.get(url_or_path, headers=HTTP_HEADERS, timeout=60)
        resp.raise_for_status()
        return resp.content
    path = url_or_path.replace("file://", "")
    with open(path, "rb") as f:
        return f.read()


def discover_latest(prefer_xlsx: bool = True) -> tuple[str | None, str | None]:
    """Scrape JADA page 342 for the most recent monthly file.

    Returns (xlsx_url, pdf_url). Either side may be None. Files are matched
    by their localised filename ("燃料別…登録台数…YYYY年M月" or similar) and
    sorted by the timestamp prefix in /files/libs/<id>/<ts><id>.<ext> so we
    pick the newest publication first.
    """
    from bs4 import BeautifulSoup

    print(f"Scanning {JADA_PAGE} for the latest monthly file …")
    resp = requests.get(JADA_PAGE, headers=HTTP_HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # Candidate links: either /relays/download/342/... (canonical) or
    # /files/libs/<id>/<ts><id>.<ext>. The relays form embeds the real file
    # path in the `file` query string and a Japanese label in `file_name`.
    candidates: list[tuple[str, str, str]] = []  # (timestamp, ext, absolute_url)

    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = (a.get_text() or "") + " " + href
        # Restrict to fuel-type registration ("燃料別 ... 登録") files.
        # We tolerate both 燃料別メーカー別登録台数 and 燃料別登録台数統計 names.
        if "燃料別" not in text and "fuel" not in text.lower():
            # The page may show only the icon link with no visible text;
            # in that case fall back to URL-only check via file_name=...
            if "file_name=" not in href:
                continue
            if "%E7%87%83%E6%96%99" not in href:  # 燃料 URL-encoded
                continue

        # Resolve to the real file path.
        if "/relays/download/342/" in href and "file=" in href:
            m = re.search(r"[?&]file=([^&]+)", href)
            if not m:
                continue
            file_path = requests.utils.unquote(m.group(1))
            url = JADA_HOST + file_path if file_path.startswith("/") else file_path
        elif "/files/libs/" in href:
            url = href if href.startswith("http") else JADA_HOST + href
        else:
            continue

        ext_m = re.search(r"\.(pdf|xlsx|xls)(?:$|\?)", url, re.IGNORECASE)
        if not ext_m:
            continue
        ext = ext_m.group(1).lower()

        ts_m = re.search(r"/files/libs/\d+/?/?(\d{8,14})", url)
        ts = ts_m.group(1) if ts_m else "00000000"

        candidates.append((ts, ext, url))

    if not candidates:
        return None, None

    # Newest first.
    candidates.sort(key=lambda x: x[0], reverse=True)

    xlsx_url = next((u for ts, ext, u in candidates if ext in ("xlsx", "xls")), None)
    pdf_url = next((u for ts, ext, u in candidates if ext == "pdf"), None)
    return xlsx_url, pdf_url


# ---------------------------------------------------------------------------
# XLSX parser
# ---------------------------------------------------------------------------

def parse_xlsx(xlsx_bytes: bytes, target_year: int, target_month: int) -> dict | None:
    """Extract the 乗用車計 row for (year, month) from an XLSX rollup.

    Returns a dict with fuel keys (PETROL/HEV/PHEV/DIESEL/BEV/OTHERS/TOTAL)
    or None if the target month is not present.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Identify month from the sheet's title cell (more reliable than the
        # sheet name in case JADA changes the latter). Scan first ~5 rows.
        title_year = title_month = None
        for row in ws.iter_rows(values_only=True, max_row=5):
            for cell in row:
                if isinstance(cell, str):
                    m = TITLE_RE.search(cell)
                    if m:
                        title_year, title_month = int(m.group(1)), int(m.group(2))
                        break
            if title_year is not None:
                break

        if title_year != target_year or title_month != target_month:
            continue

        # Find the row whose first non-empty cell == 乗用車計
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == TOTAL_ROW_LABEL:
                return _xlsx_total_row_to_fuels(row)

        # Title matched but no total row → corrupt file.
        raise RuntimeError(
            f"Sheet for {target_year}-{target_month:02d} present but missing "
            f"'{TOTAL_ROW_LABEL}' row"
        )

    return None


def _xlsx_total_row_to_fuels(row: tuple) -> dict:
    """Map the 乗用車計 row to our fuel dict.

    Column layout observed across the 2026 sample (1-indexed):
        A(1)  = '乗用車計'
        D(4)  = ガソリン   (PETROL)
        F(6)  = ＨＶ       (HEV)
        H(8)  = ＰＨＶ     (PHEV)
        J(10) = ディーゼル (DIESEL)
        L(12) = ＥＶ       (BEV)
        N(14) = ＦＣＶ     (FCV)
        P(16) = その他(*) (LPG etc.)
        R(18) = 合計       (TOTAL)

    Columns between (E,G,I,K,M,O,Q) hold 前年比 (year-on-year %) which we ignore.
    """
    # Convert to 1-based access by prepending a None.
    cells = (None,) + row

    def n(idx: int) -> int:
        v = cells[idx] if idx < len(cells) else 0
        if v is None or v == "":
            return 0
        if isinstance(v, str):
            v = v.replace(",", "").strip()
            if v in ("", "-", "--"):
                return 0
            return int(float(v))
        return int(v)

    petrol = n(4)
    hev = n(6)
    phev = n(8)
    diesel = n(10)
    bev = n(12)
    fcv = n(14)
    other = n(16)
    total = n(18)

    return {
        "PETROL": petrol,
        "HEV": hev,
        "PHEV": phev,
        "DIESEL": diesel,
        "BEV": bev,
        "OTHERS": fcv + other,
        "TOTAL": total,
    }


# ---------------------------------------------------------------------------
# PDF parser (fallback)
# ---------------------------------------------------------------------------

# A "data token" on a row is either an integer (possibly comma-grouped) or a
# percentage like "64.3" / "▲1.0" / "--". We parse a row by tokenising and
# treating tokens that look like an integer as the fuel-count columns and
# percent-like tokens as the year-on-year column between them.
_TOKEN_RE = re.compile(r"\d{1,3}(?:,\d{3})*(?:\.\d+)?|[▲△▼\-]\d+(?:\.\d+)?|--")
_INT_ONLY_RE = re.compile(r"^\d{1,3}(?:,\d{3})*$")  # no decimal point


def _parse_data_row(line: str) -> list[int] | None:
    """Split a row text into its integer fuel-count columns.

    Each fuel column is followed by a single YoY percentage token. The fuel
    counts are the tokens at even positions (0, 2, 4, …). We return the list
    of integer counts; ``None`` if the line isn't a recognisable data row
    (no commas anywhere → likely a 構成比 % row, or no tokens at all).
    """
    tokens = _TOKEN_RE.findall(line)
    # Filter out very short rows (need ≥ 3 token pairs = 2 categories + TOTAL).
    if len(tokens) < 6 or len(tokens) % 2 != 0:
        return None
    # Take every even-indexed token (column = count, not %).
    counts_raw = tokens[0::2]
    # A genuine data row has thousand-separated counts; a 構成比 row has only
    # small floats. Require at least one comma-formatted integer.
    if not any("," in c for c in counts_raw):
        return None
    counts: list[int] = []
    for c in counts_raw:
        if not _INT_ONLY_RE.match(c):
            return None  # row contained a decimal among the count positions
        counts.append(int(c.replace(",", "")))
    return counts


def parse_pdf(pdf_bytes: bytes, target_year: int, target_month: int) -> dict | None:
    """Fallback: extract 乗用車計 row from a JADA monthly PDF.

    The PDF contains one page per month (same 4-month rollup as the XLSX).
    We locate the page whose title matches target_year/month, then scan its
    lines for data rows. The 乗用車計 row is the one with the largest TOTAL
    (last column) — all manufacturer rows have smaller totals.

    Each PDF data row contains the all-makes line with 5-7 fuel-type columns
    plus a TOTAL column, each immediately followed by a YoY percentage. The
    last integer on the line is TOTAL; the values before it map to:
        6 cols → PETROL, HEV, PHEV, DIESEL, BEV, その他/FCV combined
        7 cols → PETROL, HEV, PHEV, DIESEL, BEV, FCV, その他
    JADA's published layout always has PETROL/HEV/PHEV/DIESEL/BEV/FCV/その他
    in that order. When a category was zero, the cell appears as blank in
    the PDF (and in the XLSX as 0). We treat any missing trailing categories
    as zero by counting backwards from TOTAL.
    """
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(pdf_bytes))
    for page in reader.pages:
        text = page.extract_text() or ""
        m_title = TITLE_RE.search(text)
        if not m_title:
            continue
        if int(m_title.group(1)) != target_year or int(m_title.group(2)) != target_month:
            continue

        data_rows: list[list[int]] = []
        for line in text.split("\n"):
            counts = _parse_data_row(line)
            if counts is not None:
                data_rows.append(counts)
        if not data_rows:
            raise RuntimeError(
                f"PDF page for {target_year}-{target_month:02d} matched the title "
                f"but no data row could be parsed"
            )

        # 乗用車計 has the largest TOTAL (last column).
        best = max(data_rows, key=lambda c: c[-1])
        total = best[-1]
        fuels = best[:-1]  # PETROL, HEV, PHEV, DIESEL, BEV[, FCV[, その他]]

        # Pad missing trailing categories with zeros (= unreported cells).
        while len(fuels) < 7:
            fuels.append(0)
        petrol, hev, phev, diesel, bev, fcv, other = fuels[:7]

        return {
            "PETROL": petrol,
            "HEV": hev,
            "PHEV": phev,
            "DIESEL": diesel,
            "BEV": bev,
            "OTHERS": fcv + other,
            "TOTAL": total,
        }

    return None


# ---------------------------------------------------------------------------
# CSV upsert
# ---------------------------------------------------------------------------

def build_row(period: str, fuels: dict, source_url: str) -> dict:
    return {
        "period": period,
        "time_interval": "monthly",
        "variant": "Whole",
        "source": "JADA /JAMA",
        "BEV": float(fuels["BEV"]),
        "PHEV": float(fuels["PHEV"]),
        "HEV": float(fuels["HEV"]),
        "PETROL": float(fuels["PETROL"]),
        "DIESEL": float(fuels["DIESEL"]),
        "OTHERS": float(fuels["OTHERS"]),
        "TOTAL": float(fuels["TOTAL"]),
        "notes": source_url,
    }


def upsert_row(csv_path: str, period: str, row: dict, force: bool) -> bool:
    existing: dict[str, dict] = {}
    line_ending = "\n"
    if Path(csv_path).exists():
        # Detect the existing file's line ending so the diff stays minimal
        # when the workflow commits. data/Japan.csv is currently CRLF.
        with open(csv_path, "rb") as fb:
            head = fb.read(4096)
        if b"\r\n" in head:
            line_ending = "\r\n"
        with open(csv_path, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                existing[r["period"]] = r

    if period in existing and not force:
        print(f"  Period {period} already in CSV — not overwriting (use --force).")
        return False

    existing[period] = row
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, lineterminator=line_ending)
        writer.writeheader()
        for p in sorted(existing.keys()):
            writer.writerow(existing[p])
    return True


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--year", type=int)
    parser.add_argument("--month", type=int, choices=range(1, 13))
    parser.add_argument("--xlsx-url", help="Direct URL/path to the JADA monthly XLSX")
    parser.add_argument("--pdf-url", help="Direct URL/path to the JADA monthly PDF (fallback)")
    parser.add_argument("--csv", default="data/Japan.csv")
    parser.add_argument("--force", action="store_true",
                        help="Re-process even if target period already exists")
    args = parser.parse_args()

    if args.year and args.month:
        target_year, target_month = args.year, args.month
    elif args.year or args.month:
        sys.exit("--year and --month must be given together")
    else:
        target_year, target_month = previous_month(date.today())
    target_period = f"{target_year}-{target_month:02d}"
    print(f"Target period: {target_period}")

    # Self-throttle: skip if CSV is already at-or-past the target.
    if not args.force:
        latest = latest_period(args.csv)
        if latest and latest >= target_period:
            print(f"Latest period in CSV is {latest} ≥ {target_period} — nothing to do.")
            return 0

    xlsx_src = args.xlsx_url
    pdf_src = args.pdf_url

    if not xlsx_src and not pdf_src:
        try:
            xlsx_src, pdf_src = discover_latest()
        except requests.HTTPError as e:
            print(f"Could not reach {JADA_PAGE}: {e}. "
                  "Pass --xlsx-url or --pdf-url manually.")
            return 0
        if not xlsx_src and not pdf_src:
            print("No JADA fuel-registration file found on the index page. "
                  "Will retry on next scheduled run.")
            return 0
        if xlsx_src: print(f"Found XLSX: {xlsx_src}")
        if pdf_src: print(f"Found PDF:  {pdf_src}")

    fuels: dict | None = None
    source_url = ""

    if xlsx_src:
        try:
            fuels = parse_xlsx(load_bytes(xlsx_src), target_year, target_month)
            if fuels is not None:
                source_url = xlsx_src
        except Exception as e:
            print(f"XLSX parsing failed ({e}); will try PDF fallback.")

    if fuels is None and pdf_src:
        fuels = parse_pdf(load_bytes(pdf_src), target_year, target_month)
        if fuels is not None:
            source_url = pdf_src

    if fuels is None:
        print(f"Target month {target_period} not present in the latest JADA file. "
              "Will retry on next scheduled run.")
        return 0

    # Sanity: components sum to TOTAL.
    components = fuels["PETROL"] + fuels["HEV"] + fuels["PHEV"] + fuels["DIESEL"] + fuels["BEV"] + fuels["OTHERS"]
    if components != fuels["TOTAL"]:
        sys.exit(
            f"Sanity check failed: PETROL+HEV+PHEV+DIESEL+BEV+OTHERS={components} "
            f"!= TOTAL={fuels['TOTAL']}. Parsed values: {fuels}"
        )
    print(f"Parsed: BEV={fuels['BEV']}, PHEV={fuels['PHEV']}, HEV={fuels['HEV']}, "
          f"PETROL={fuels['PETROL']}, DIESEL={fuels['DIESEL']}, OTHERS={fuels['OTHERS']}, "
          f"TOTAL={fuels['TOTAL']}")

    row = build_row(target_period, fuels, source_url)
    if upsert_row(args.csv, target_period, row, args.force):
        print(f"\nWrote {target_period} to {args.csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
