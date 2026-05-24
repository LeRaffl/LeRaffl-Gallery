#!/usr/bin/env python3
"""
Fetch Uruguay vehicle registration data from ACAU and update data/Uruguay.csv.

Usage
-----
    python scripts/fetch_uruguay.py [--year YEAR] [--url URL_OR_PATH] \
        [--csv PATH] [--force]

* --year   Year to fetch (default: current calendar year).
* --url    Direct Compilado xlsx URL or local file path; if omitted, the
           ACAU homepage is scraped for the year's Compilado link.
* --csv    Target CSV (default: data/Uruguay.csv).
* --force  Re-process months even if already present in the CSV.

Invoked by .github/workflows/fetch-uruguay.yml on a daily cron from the 1st
of each month onward, plus manual workflow_dispatch. When the CSV changes,
the workflow commits data/Uruguay.csv and triggers render-country.yml for
Uruguay.

Data source
-----------
ACAU (Asociación del Comercio Automotor del Uruguay) publishes two xlsx
workbooks per year on https://www.acau.com.uy/ :

  1. "Compilado YYYY"  — per-model rows, one sheet per vehicle category
     (AUTOS, SUV, MINIBUSES, UTILITARIO, CAMIONES, OMNIBUS), monthly volumes
     plus a year-to-date total column. **We parse this one** — specifically
     the AUTOS and SUV sheets, summed together.
  2. "Mercado YYYY"    — per-manufacturer monthly totals. Not currently
     ingested; useful as a maintainer-side cross-check only.

The download filenames look like ``15_18_25ar1.xlsx`` (HH_MM_SS of the local
Uruguay upload time + "ar1.xlsx") — i.e. unpredictable, so we have to scrape
the homepage to discover the current year's link rather than hard-coding it.

Vehicle scope
-------------
AUTOS (turismos: sedans, hatchbacks, coupés) + SUV (utility vehicles). The
ACAU spreadsheet uses these as two separate sheets but the maintainer wants
them aggregated into a single passenger-car series. MINIBUSES, UTILITARIO
(light commercial / pickups), CAMIONES (medium/heavy trucks) and OMNIBUS
(buses) are explicitly out of scope — Uruguay has no HDV variant yet and
the light-commercial / minibus categories don't map cleanly onto the
project's existing CSV schemas. See docs/architecture/09-glossary.md
§ Vehicle scope per source.

Fuel mapping (ACAU code → CSV column)
-------------------------------------
    E    (Eléctrico)        → BEV
    PHEV (Plug-in híbrido)  → PHEV
    H    (Híbrido)          → HEV   (full / regular hybrids)
    N    (Nafta)            → PETROL
    D    (Diesel)           → DIESEL
    MHEV (Mild hybrid)      → OTHERS

MHEV → OTHERS: per maintainer call, mild hybrids are bucketed into OTHERS
rather than mixed into HEV (which would inflate full-hybrid counts) or
PETROL/DIESEL (where they technically belong, but losing the distinction
makes the EV-share series jittery whenever a manufacturer relabels a model).
Same convention isn't used for Chile because Chile's CSV has no OTHERS
column the way other sources do — there MHEV falls into ICE via subtraction.

Pre-2026 files (e.g. 2025) have a different sheet/column layout (sheet
"MINI" instead of "MINIBUSES", per-brand subtotal rows, abbreviated month
headers "Ene"/"Feb"/…) and didn't break out PHEV at all — the maintainer
had to google models to classify them. This parser targets the 2026+ layout
only; back-filling pre-2026 years would require a separate parser.

CSV layout
----------
    period,time_interval,variant,source,BEV,PHEV,HEV,PETROL,DIESEL,OTHERS,TOTAL,notes

Months without published data
-----------------------------
ACAU pre-fills the entire calendar year with zeros and overwrites
month-by-month as data becomes available. We skip any month where ALL fuel
values across AUTOS + SUV are zero — this lets the script run mid-year
without producing fake "all-zero" rows for future months.

Upsert + sanity check
---------------------
Each parsed month writes a CSV row keyed by ``period`` (YYYY-MM). New
periods are appended; already-present periods are skipped unless --force.
The bottom "TOTAL" row of each sheet is read and cross-checked against
our per-fuel sum — if they don't match for a published month, the parser
fails loudly (likely indicates a layout change we haven't handled).

HTTP details
------------
ACAU runs LiteSpeed and didn't reject python-requests with a default UA in
testing, but we send a desktop-Chrome UA + Spanish Accept-Language anyway
for consistency with the other fetchers and to be a polite robot.
"""
import argparse
import csv
import io
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup

ACAU_HOME = "https://www.acau.com.uy/"
ACAU_HOST = "https://www.acau.com.uy"

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-UY,es;q=0.9,en;q=0.8",
}

CSV_COLUMNS = [
    "period", "time_interval", "variant", "source",
    "BEV", "PHEV", "HEV", "PETROL", "DIESEL", "OTHERS",
    "TOTAL", "notes",
]

# AUTOS + SUV only — see docstring "Vehicle scope" for why.
TARGET_SHEETS = ("AUTOS", "SUV")

# Combustible code → CSV column. Lookup is case-insensitive (the source
# mixes upper and lower case: 'N'/'n', 'D'/'d', etc.).
FUEL_MAP = {
    "E":    "BEV",
    "PHEV": "PHEV",
    "H":    "HEV",
    "N":    "PETROL",
    "D":    "DIESEL",
    "MHEV": "OTHERS",
}

SPANISH_MONTHS = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def find_compilado_url(year: int) -> str:
    """Scrape the ACAU homepage and return the Compilado xlsx URL for `year`.

    The page lists each year's Compilado/Mercado pair under
    ``div.item_estadistica`` blocks; the link text is e.g. "Compilado 2026".
    Filenames are unpredictable (timestamp-based), so we match on the text
    label rather than the URL.
    """
    print(f"Scanning {ACAU_HOME} for Compilado {year} …")
    resp = requests.get(ACAU_HOME, headers=HTTP_HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    label_re = re.compile(rf"\bCompilado\s+{year}\b", re.IGNORECASE)
    for a in soup.find_all("a", href=True):
        if not a["href"].lower().endswith(".xlsx"):
            continue
        if label_re.search(a.get_text(" ", strip=True)):
            href = a["href"]
            if href.startswith("http"):
                return href
            # Relative paths on this site look like "../panel/estadisticas/<f>.xlsx"
            # — resolve against the page root.
            return ACAU_HOST + "/" + href.lstrip("./")

    raise RuntimeError(
        f"Could not find a Compilado {year} link on {ACAU_HOME}. "
        "Pass --url to specify the xlsx directly."
    )


def load_bytes(url_or_path: str) -> bytes:
    if url_or_path.startswith("http://") or url_or_path.startswith("https://"):
        print(f"Downloading: {url_or_path}")
        resp = requests.get(url_or_path, headers=HTTP_HEADERS, timeout=60)
        resp.raise_for_status()
        return resp.content
    path = url_or_path.replace("file://", "")
    with open(path, "rb") as f:
        return f.read()


def parse_sheet(ws) -> tuple[dict[str, list[float]], list[float] | None]:
    """Walk one sheet (AUTOS or SUV) and return per-month fuel counts.

    Returns ``({csv_column: [12 monthly floats]}, total_per_month_or_None)``,
    where ``total_per_month`` is the 12 values pulled from the sheet's
    bottom "TOTAL" row (in calendar-month order Enero..Diciembre) when one
    is present, used by the caller as a sanity check.

    Layout assumption (2026):
        row 5: "COMPILADO YYYY"            (year header — checked by caller)
        row 6: sheet kind ("AUTOMOVILES" / "S.U.V.")
        row 7: blank
        row 8: column headers — first cell is "Nombre_Socio", contains
               "Combustible" plus 12 month names "Enero" … "Diciembre"
        row 9+: data rows; one per model
        last:  "TOTAL" in column A with monthly totals (used as sanity check)

    The header row is located dynamically by looking for the cell value
    "Combustible" so the parser tolerates small column shuffles. Month
    columns are matched against ``SPANISH_MONTHS`` so any abbreviated /
    reordered variant fails loudly rather than silently shifting columns.
    """
    rows = list(ws.iter_rows(values_only=True))

    # Locate header row by finding the cell that says "Combustible".
    hdr_idx = None
    for i, row in enumerate(rows[:25]):
        if any(isinstance(c, str) and c.strip() == "Combustible" for c in row):
            hdr_idx = i
            break
    if hdr_idx is None:
        raise RuntimeError(f"Could not locate 'Combustible' header in sheet '{ws.title}'")

    header = rows[hdr_idx]
    try:
        fuel_col = next(i for i, c in enumerate(header)
                        if isinstance(c, str) and c.strip() == "Combustible")
    except StopIteration:
        raise RuntimeError(f"'Combustible' column not found in header of '{ws.title}'")

    # Month columns: locate every Spanish month name in the header.
    month_to_col: dict[str, int] = {}
    for i, c in enumerate(header):
        if isinstance(c, str) and c.strip() in SPANISH_MONTHS:
            month_to_col[c.strip()] = i
    missing = [m for m in SPANISH_MONTHS if m not in month_to_col]
    if missing:
        raise RuntimeError(
            f"Sheet '{ws.title}' header is missing month columns {missing}. "
            "Layout may have changed."
        )

    # Aggregate per-fuel monthly counts. Unknown fuel codes are surfaced as
    # warnings but their counts are still folded into OTHERS so the per-month
    # sanity check still balances against the file's own TOTAL row.
    fuel_totals: dict[str, list[float]] = {col: [0.0] * 12 for col in set(FUEL_MAP.values())}
    total_row = None
    unknown_codes: set[str] = set()

    for row in rows[hdr_idx + 1:]:
        first = row[0]
        # End-of-sheet TOTAL row.
        if isinstance(first, str) and first.strip().upper() == "TOTAL":
            total_row = row
            break

        code_raw = row[fuel_col] if fuel_col < len(row) else None
        if code_raw is None or (isinstance(code_raw, str) and code_raw.strip() == ""):
            # Blank Combustible — separator / subtotal / blank row. Skip.
            continue

        code = str(code_raw).strip().upper()
        csv_col = FUEL_MAP.get(code)
        if csv_col is None:
            unknown_codes.add(code)
            csv_col = "OTHERS"

        for j, month in enumerate(SPANISH_MONTHS):
            col = month_to_col[month]
            v = row[col] if col < len(row) else None
            if isinstance(v, (int, float)):
                fuel_totals[csv_col][j] += float(v)

    if unknown_codes:
        print(f"  WARNING {ws.title}: unknown Combustible codes folded into OTHERS: "
              f"{sorted(unknown_codes)}")

    # Extract the 12 monthly totals from the TOTAL row (if found) in calendar order.
    total_per_month = None
    if total_row is not None:
        total_per_month = []
        for month in SPANISH_MONTHS:
            col = month_to_col[month]
            v = total_row[col] if col < len(total_row) else None
            total_per_month.append(float(v) if isinstance(v, (int, float)) else 0.0)

    return fuel_totals, total_per_month


def parse_workbook(wb_bytes: bytes, year: int) -> dict[str, dict]:
    """Parse AUTOS + SUV sheets into per-period CSV rows.

    Returns ``{period: row_dict}`` for every month that has at least one
    non-zero fuel value. The workbook's ``COMPILADO YYYY`` header is
    cross-checked against ``year`` — mismatch is a hard fail.
    """
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)

    # The "COMPILADO YYYY" cell lives near the top of every sheet; pick the
    # first sheet we'll parse and validate the year matches the request.
    first = wb[TARGET_SHEETS[0]]
    year_cell = None
    for row in first.iter_rows(values_only=True, max_row=10):
        for c in row:
            if isinstance(c, str) and "COMPILADO" in c.upper():
                m = re.search(r"\b(20\d{2})\b", c)
                if m:
                    year_cell = int(m.group(1))
                    break
        if year_cell:
            break
    if year_cell is None:
        raise RuntimeError(
            "Could not find 'COMPILADO YYYY' header — file may not be a Compilado xlsx"
        )
    if year_cell != year:
        raise RuntimeError(
            f"Workbook header says COMPILADO {year_cell} but --year is {year}. "
            "Refusing to write mismatched data."
        )

    # Sum AUTOS + SUV.
    combined: dict[str, list[float]] = {col: [0.0] * 12 for col in set(FUEL_MAP.values())}
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(
                f"Workbook is missing required sheet '{sheet_name}'. "
                f"Sheets present: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        per_fuel, sheet_total = parse_sheet(ws)

        # Cross-check this sheet's per-fuel sum against its own TOTAL row.
        if sheet_total is not None:
            computed = [sum(per_fuel[f][j] for f in per_fuel) for j in range(12)]
            if computed != sheet_total:
                diffs = [(SPANISH_MONTHS[j], computed[j], sheet_total[j])
                         for j in range(12) if computed[j] != sheet_total[j]]
                raise RuntimeError(
                    f"Sheet '{sheet_name}' fuel sum != TOTAL row for months {diffs}. "
                    "Parser likely missed rows."
                )
            print(f"  {sheet_name}: per-month totals OK ({sheet_total})")
        else:
            print(f"  NOTE {sheet_name}: no TOTAL row found, skipping per-sheet sanity check")

        for f, vals in per_fuel.items():
            for j in range(12):
                combined[f][j] += vals[j]

    # Build per-month CSV rows; skip months where all fuel values are zero
    # (future months — ACAU pre-fills the year with zeros).
    rows: dict[str, dict] = {}
    for j in range(12):
        values = {f: combined[f][j] for f in combined}
        if all(v == 0.0 for v in values.values()):
            continue

        period = f"{year}-{j + 1:02d}"
        row = {
            "period": period,
            "time_interval": "monthly",
            "variant": "Whole",
            "source": "ACAU",
            "BEV":    values.get("BEV",    0.0),
            "PHEV":   values.get("PHEV",   0.0),
            "HEV":    values.get("HEV",    0.0),
            "PETROL": values.get("PETROL", 0.0),
            "DIESEL": values.get("DIESEL", 0.0),
            "OTHERS": values.get("OTHERS", 0.0),
        }
        row["TOTAL"] = sum(row[c] for c in ["BEV", "PHEV", "HEV", "PETROL", "DIESEL", "OTHERS"])
        rows[period] = row

    return rows


def upsert_csv(csv_path: str, new_rows: dict[str, dict], source_url: str,
               force: bool) -> tuple[int, int]:
    """Upsert new_rows by period. Returns (added, updated_with_force) counts.

    Without --force, periods already in the CSV are left alone (rule of
    thumb in this project: don't silently re-write historical rows). With
    --force, every period in new_rows is overwritten.

    The existing file's line-ending convention is detected and preserved
    (data/Uruguay.csv on disk is CRLF; rewriting it as LF on the first
    ingest would otherwise produce a noisy diff of every row — same
    gotcha as data/Japan.csv).
    """
    existing: dict[str, dict] = {}
    line_ending = "\n"
    if Path(csv_path).exists():
        with open(csv_path, "rb") as fb:
            head = fb.read(4096)
        if b"\r\n" in head:
            line_ending = "\r\n"
        with open(csv_path, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                existing[r["period"]] = r

    added = updated = 0
    for period, new_row in sorted(new_rows.items()):
        new_row["notes"] = source_url
        if period not in existing:
            existing[period] = new_row
            added += 1
            print(f"  + {period}")
        elif force:
            existing[period] = {**existing[period], **new_row}
            updated += 1
            print(f"  ~ {period} (--force)")
        else:
            print(f"  = {period} (already present, skipping)")

    if added == 0 and updated == 0:
        return 0, 0

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, lineterminator=line_ending)
        writer.writeheader()
        for p in sorted(existing.keys()):
            writer.writerow(existing[p])
    return added, updated


def latest_period(csv_path: str) -> str | None:
    if not Path(csv_path).exists():
        return None
    with open(csv_path, newline="", encoding="utf-8") as f:
        periods = [row["period"] for row in csv.DictReader(f)]
    return max(periods) if periods else None


def previous_month(today: date) -> tuple[int, int]:
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--year", type=int,
                        help="Compilado year to fetch (default: previous month's year)")
    parser.add_argument("--url", help="Direct Compilado xlsx URL or local path")
    parser.add_argument("--csv", default="data/Uruguay.csv")
    parser.add_argument("--force", action="store_true",
                        help="Re-process periods that already exist in the CSV")
    args = parser.parse_args()

    # Target period drives the self-throttle: we only run when the previous
    # calendar month is not yet in the CSV (matching the maintainer's
    # "schauen ob der Vormonat von Heute noch nicht in den Daten is").
    today = date.today()
    target_year, target_month = previous_month(today)
    target_period = f"{target_year}-{target_month:02d}"
    year = args.year or target_year
    print(f"Today: {today}. Target month: {target_period}. Fetching Compilado {year}.")

    if not args.force:
        latest = latest_period(args.csv)
        if latest and latest >= target_period:
            print(f"Latest period in CSV is {latest} ≥ {target_period} — nothing to do.")
            return 0

    url = args.url
    if not url:
        try:
            url = find_compilado_url(year)
            print(f"Found: {url}")
        except (requests.HTTPError, requests.ConnectionError) as e:
            print(f"Could not reach {ACAU_HOME}: {e}. "
                  "Pass --url manually or retry later.")
            return 0

    xlsx_bytes = load_bytes(url)
    new_rows = parse_workbook(xlsx_bytes, year)
    if not new_rows:
        print("No published months found in the workbook (all-zero or empty). "
              "Will retry on next scheduled run.")
        return 0
    print(f"Parsed {len(new_rows)} months: {sorted(new_rows.keys())}")

    if target_period not in new_rows and not args.force:
        # The Compilado file is published but hasn't been refreshed yet
        # with the previous month's data. Bail without writing — we'll
        # retry tomorrow.
        print(f"Target month {target_period} not yet present in the file "
              f"(latest published: {max(new_rows.keys())}). Will retry tomorrow.")
        return 0

    added, updated = upsert_csv(args.csv, new_rows, url, args.force)
    print(f"\nDone: {added} rows added, {updated} rows updated → {args.csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
