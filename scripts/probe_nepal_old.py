#!/usr/bin/env python3
"""
TEMPORARY probe for the Nepal pre-2020 backfill investigation.

Discovers the old fiscal-year FTS categories, downloads their workbooks, and
writes a COMPACT text report (sheet names + header rows + a sample of 8703
rows for every import-looking sheet) to scratch_nepal/oldfmt_report.txt so
the old layouts can be analysed locally without committing the raw files.

Removed again once the backfill decision is made.
"""
import io
import sys
from pathlib import Path

import openpyxl
import requests

sys.path.insert(0, str(Path(__file__).resolve().parent))
import fetch_nepal as fn  # noqa: E402

FYS = [2076, 2075, 2074, 2073, 2072, 2071]
OUT = Path("scratch_nepal/oldfmt_report.txt")
IMPORT_HINT = ("import", "hswise", "commodit", "hs")


def dump_sheet(ws, name, lines):
    lines.append(f"    -- sheet {name!r}  dims={ws.max_row}x{ws.max_column}")
    rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
    for i, row in enumerate(rows):
        cells = ["" if c is None else str(c)[:26] for c in (row or [])][:8]
        lines.append(f"       r{i}: {cells}")
    # count + sample 8703 rows anywhere in the sheet
    n8703, samples = 0, []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        c0 = "" if row[0] is None else str(row[0]).strip()
        if c0.startswith("8703"):
            n8703 += 1
            if len(samples) < 6:
                samples.append(["" if c is None else str(c)[:30] for c in row][:6])
    lines.append(f"       8703 rows: {n8703}")
    for s in samples:
        lines.append(f"         {s}")


def main() -> int:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers.update(fn.HTTP_HEADERS)
    cats = fn.discover_fy_categories(session)
    lines: list[str] = []
    for fy in FYS:
        lines.append(f"\n===== FY {fy}/{(fy+1)%100:02d}")
        if fy not in cats:
            lines.append("  (no category)")
            continue
        content = fn.find_fts_content_page(session, fy, cats[fy])
        if content is None:
            lines.append("  (no FTS content item)")
            continue
        try:
            wbs = fn.list_workbooks(session, content)
        except Exception as exc:  # noqa: BLE001
            lines.append(f"  list_workbooks failed: {exc}")
            continue
        lines.append(f"  content={content}  ({len(wbs)} workbooks)")
        # first 3 workbooks per FY is enough to characterise the layout
        for url, label in wbs[:3]:
            fname = url.rsplit("/", 1)[-1][:50]
            lines.append(f"  FILE {label!r}: {fname}")
            try:
                r = session.get(url, timeout=180)
                r.raise_for_status()
                wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
            except Exception as exc:  # noqa: BLE001
                lines.append(f"    download/open failed: {exc}")
                continue
            lines.append(f"    sheets: {wb.sheetnames}")
            for sn in wb.sheetnames:
                if any(h in sn.lower() for h in IMPORT_HINT):
                    try:
                        dump_sheet(wb[sn], sn, lines)
                    except Exception as exc:  # noqa: BLE001
                        lines.append(f"    -- sheet {sn!r} dump failed: {exc}")
    OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"wrote {OUT} ({len(lines)} lines)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
