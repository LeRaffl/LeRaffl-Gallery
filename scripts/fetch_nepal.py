#!/usr/bin/env python3
"""
Fetch Nepal vehicle-import data from the Department of Customs' monthly
Foreign Trade Statistics (FTS) workbooks and upsert data/Nepal.csv plus
data/Nepal_3-Wheelers.csv.

Usage
-----
    python scripts/fetch_nepal.py                  # current FY (steady state)
    python scripts/fetch_nepal.py --fy 2081        # one specific FY (BS start year)
    python scripts/fetch_nepal.py --backfill       # every FY found on the site
    python scripts/fetch_nepal.py --from-dir DIR   # offline: parse local .xlsx files
    python scripts/fetch_nepal.py --force          # rewrite even if nothing changed

Source & semantics
------------------
The Department of Customs (customs.gov.np) publishes one cumulative
fiscal-year-to-date FTS workbook per Nepali month (sheet
"5_Imports_By_Commodity": 8-digit HS code, unit, quantity). Nepal has no
domestic car manufacturing, so **imports ≈ the national new-vehicle market**
(the same source every Nepali press report quotes). Figures are imports,
not registrations — same caveat class as Indonesia (wholesales) / China
(Wholesale variant).

Two series are produced from heading 87.03 (motor cars & other motor
vehicles principally designed for the transport of persons — ≈ EU M1 plus
Nepal's passenger three-wheelers, which we split out):

    Whole       data/Nepal.csv              8703 minus three-wheelers
    3-Wheelers  data/Nepal_3-Wheelers.csv   auto-rickshaws + e-rickshaws
                                            (87032111/19, 87038011/19, …)

Out of scope (documented in docs/architecture/32-source-nepal.md): buses &
10+-seat vans (8702, incl. the large electric mini/микrobus wave), goods
vehicles/pickups (8704), motorcycles (8711).

Fuel mapping (by 6-digit HS prefix; description only decides the
three-wheeler split and an electric override for legacy codes):

    870310            -> OTHERS   (golf cars etc.)
    870321..870324    -> PETROL
    870331..870333    -> DIESEL
    870340, 870350    -> HEV  \\  single "Hybrid bucket" convention
    870360, 870370    -> HEV  /   (see below)
    870380            -> BEV
    870390            -> OTHERS (desc containing "electric" -> BEV)

**Single Hybrid bucket:** DoC's own 8-digit descriptions split hybrids by
engine displacement, not by plug-in capability ("plugin UPTO 2000CC" on
8703.40, "plug out ABOVE 2000 CC" on 8703.60 — mutually contradictory and
inconsistent with the international 6-digit meaning). The PHEV/HEV boundary
in this source is therefore not trustworthy, so all hybrids go into the HEV
column with PHEV left empty — the Türkiye/Georgia/Colombia convention; the
charts label the bucket "Hybrid". Volumes are tiny either way (85 units in
FY 2081/82 vs ~13.6k BEVs).

Calendar
--------
Nepali fiscal year = Shrawan 1 .. Ashadh end (~17 Jul .. ~16 Jul). FY
"2082/83" starts in AD year 2082-57 = 2025. Each fiscal month k (1..12) is
labelled with the Gregorian month in which the Nepali month *ends*:

    period(fy_start, k) = (July of AD fy_start-57) + k months
    e.g. FY2082/83: Shrawan -> 2025-08, Poush -> 2026-01, Ashadh -> 2026-07

This fixed rule gives a gapless, duplicate-free monthly series across FY
boundaries (Nepali months straddle two Gregorian months; the label month
contains the month's end and roughly half its days).

Monthly values
--------------
Each workbook is cumulative FYTD, so month k = FYTD(k) - FYTD(k-1)
(Shrawan = its own file). Every run re-downloads all published files of the
processed FY and rebuilds all its months, so upstream revisions are
absorbed (Indonesia-style). Coverage (FY + month count) is parsed from the
workbook's own English header line ("Based on First Eleven Months
(Shrawan-Jestha) of FY 2081/82 ...") — the Devanagari link labels and
filenames on the site are too inconsistent to trust ("जेठसम्म" vs
"जेष्ठसम्म", "आ.ब." vs "आ.व.", files named "फागुनसम्म_nwsjsgt.xlsx").

Validation (hard fail, no partial writes)
-----------------------------------------
* header line must parse to (fy, month_count) and the FY must match the
  content page it was discovered on;
* month set of a FY must be contiguous 1..K (deltas need every step);
* negative monthly deltas abort (would mean DoC revised a cumulative
  figure downward — needs a human look);
* every parsed 8703 row must carry unit PCS (warned otherwise) and a
  numeric quantity;
* unknown 8703 sub-codes are bucketed into OTHERS with a loud warning.

Discovery (three hops, all server-rendered HTML)
------------------------------------------------
1. homepage nav -> FTS fiscal-year category links: anchors titled
   "आ.व. <BS>/<BS+1>" (category slugs are NOT predictable: fts-2081-082,
   a-v-2042-063, ...). Both the FTS dropdown and a revenue-summary
   dropdown use the same title pattern, so every candidate category is
   fetched and disambiguated in hop 2.
2. category page -> the content item whose title contains "तथ्याङ्क"
   (statistics) and not "राजस्व" (revenue). Categories may be empty
   (pre-2080/81 FYs were left behind on archive.customs.gov.np).
3. content page -> <div class="details__desc"> holds one <a> per
   published workbook (giwmscdnone.gov.np/media/files/*.xlsx).

See docs/architecture/32-source-nepal.md for the full playbook.
"""
import argparse
import csv
import io
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
import requests

BASE = "https://www.customs.gov.np"
SOURCE = "customs.gov.np (Department of Customs FTS, HS 8703 imports)"

CSV_COLUMNS = [
    "period", "time_interval", "variant", "source",
    "BEV", "PHEV", "HEV", "PETROL", "DIESEL", "OTHERS",
    "TOTAL", "notes",
]

VARIANT_CONFIG = {
    "Whole":      "data/Nepal.csv",
    "3-Wheelers": "data/Nepal_3-Wheelers.csv",
}

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    ),
    "Accept": "*/*",
}

# Devanagari digits -> ASCII
DEVANAGARI_DIGITS = str.maketrans("०१२३४५६७८९", "0123456789")

# fiscal-month index (Shrawan=1 .. Ashadh=12) by English month name as it
# appears in the workbook header's "(Shrawan-Jestha)" range. Spellings vary
# ("Push"/"Poush", "Asoj"/"Ashwin"), so match loosely and lowercase.
FY_MONTH_BY_NAME = {
    "shrawan": 1, "shravan": 1, "sawan": 1, "saun": 1,
    "bhadra": 2, "bhadau": 2,
    "asoj": 3, "ashwin": 3, "aswin": 3, "asauj": 3,
    "kartik": 4, "karttik": 4,
    "mangsir": 5, "mangshir": 5, "marga": 5,
    "push": 6, "poush": 6, "push(poush)": 6, "pus": 6, "paush": 6,
    "magh": 7,
    "falgun": 8, "fagun": 8, "phalgun": 8,
    "chaitra": 9, "chait": 9,
    "baishakh": 10, "baisakh": 10, "baishak": 10, "vaishakh": 10,
    "jestha": 11, "jeth": 11, "jyestha": 11,
    "ashadh": 12, "ashad": 12, "asar": 12, "asadh": 12, "aashad": 12,
}

ORDINAL_MONTH_COUNT = {
    "first": 1, "one": 1,
    "second": 2, "two": 2,
    "third": 3, "three": 3,
    "fourth": 4, "four": 4,
    "fifth": 5, "five": 5,
    "sixth": 6, "six": 6,
    "seventh": 7, "seven": 7,
    "eighth": 8, "eight": 8,
    "ninth": 9, "nine": 9,
    "tenth": 10, "ten": 10,
    "eleventh": 11, "eleven": 11,
    "twelfth": 12, "twelve": 12,
}

THREE_WHEELER_RE = re.compile(r"rik?shaw|three\s*wheel|3\s*wheel", re.I)
ELECTRIC_RE = re.compile(r"electric", re.I)

# 6-digit HS prefix -> fuel bucket (within heading 8703).
PREFIX_BUCKET = {
    "870310": "OTHERS",
    "870321": "PETROL", "870322": "PETROL", "870323": "PETROL", "870324": "PETROL",
    "870331": "DIESEL", "870332": "DIESEL", "870333": "DIESEL",
    "870340": "HEV", "870350": "HEV",
    # single Hybrid bucket: DoC's plug-in/non-plug-in split is CC-garbled,
    # so 8703.60/.70 join 8703.40/.50 in HEV (labelled "Hybrid" downstream).
    "870360": "HEV", "870370": "HEV",
    "870380": "BEV",
    "870390": "OTHERS",
}

BUCKETS = ("BEV", "HEV", "PETROL", "DIESEL", "OTHERS")


# --------------------------------------------------------------------------
# Calendar
# --------------------------------------------------------------------------

def period_for(bs_fy_start: int, fy_month: int) -> str:
    """Gregorian YYYY-MM label for fiscal month 1..12 of FY <bs_fy_start>/<+1>.

    Label = the Gregorian month in which the Nepali month ends (July of the
    AD start year + fy_month). Gapless across FY boundaries:
    FY2081 Ashadh -> 2025-07, FY2082 Shrawan -> 2025-08.
    """
    ad = bs_fy_start - 57
    m = 7 + fy_month
    return f"{ad + (m - 1) // 12:04d}-{(m - 1) % 12 + 1:02d}"


# --------------------------------------------------------------------------
# Discovery
# --------------------------------------------------------------------------

FY_TITLE_RE = re.compile(r"आ\.\s*[वब]\.?\s*([०-९0-9]{4})\s*[/-]")
ANCHOR_RE = re.compile(r'<a[^>]*href="([^"#]+)"[^>]*>([\s\S]*?)</a>')
MEDIA_XLSX_RE = re.compile(
    r'<a[^>]*href="(https://[a-z0-9.]+\.gov\.np/media/files/[^"]+\.xlsx)"[^>]*>([\s\S]*?)</a>',
    re.I,
)


def anchor_text(raw: str) -> str:
    txt = re.sub(r"<[^>]+>", "", raw)
    txt = txt.replace("&nbsp;", " ")
    return re.sub(r"\s+", " ", txt).strip()


def get(session: requests.Session, url: str) -> str:
    r = session.get(url, timeout=90)
    r.raise_for_status()
    return r.text


def discover_fy_categories(session: requests.Session) -> dict[int, list[str]]:
    """Map BS fiscal start year -> candidate category URLs from the homepage nav."""
    html = get(session, BASE + "/")
    out: dict[int, list[str]] = {}
    for href, raw in ANCHOR_RE.findall(html):
        title = anchor_text(raw)
        m = FY_TITLE_RE.search(title)
        if not m or "/category/" not in href:
            continue
        year = int(m.group(1).translate(DEVANAGARI_DIGITS))
        url = href.strip()
        if url.startswith("/"):
            url = BASE + url
        out.setdefault(year, [])
        if url not in out[year]:
            out[year].append(url)
    if not out:
        raise RuntimeError(
            "no 'आ.व. YYYY/YY' category links found on the customs.gov.np "
            "homepage — nav layout may have changed"
        )
    return out


def find_fts_content_page(session: requests.Session, fy: int,
                          category_urls: list[str]) -> str | None:
    """Return the FTS statistics content-page URL inside one of the FY's
    candidate categories, or None if the FY has no statistics item (empty
    category / revenue-summary-only category)."""
    for cat_url in category_urls:
        html = get(session, cat_url)
        for href, raw in ANCHOR_RE.findall(html):
            title = anchor_text(raw)
            if "/content/" not in href or not title:
                continue
            # reject the revenue summary, which shares the आ.व. title pattern
            if "राजस्व" in title:
                continue
            # "तथ्याङ्क"/"तथ्यांक" = statistics (spelling varies). Some FY
            # items are titled with the bare fiscal year only ("आ.व.२०८०/०८१"),
            # so an आ.व. title whose year matches this category's FY counts
            # too. A wrong pick still can't corrupt data: every workbook's
            # header FY is verified against `fy` before parsing.
            is_stats = "तथ्याङ्क" in title or "तथ्यांक" in title
            m = FY_TITLE_RE.search(title)
            is_fy_item = bool(m) and int(m.group(1).translate(DEVANAGARI_DIGITS)) == fy
            if is_stats or is_fy_item:
                url = href.strip()
                return BASE + url if url.startswith("/") else url
    return None


def list_workbooks(session: requests.Session, content_url: str) -> list[tuple[str, str]]:
    """Return [(xlsx_url, link_label), ...] from a FTS content page."""
    html = get(session, content_url)
    pairs = [(u, anchor_text(t)) for u, t in MEDIA_XLSX_RE.findall(html)]
    if not pairs:
        raise RuntimeError(f"no .xlsx attachments found on {content_url}")
    return pairs


# --------------------------------------------------------------------------
# Workbook parsing
# --------------------------------------------------------------------------

HEADER_FY_RE = re.compile(r"F\.?Y\.?\s*(\d{4})\s*[/-]", re.I)
HEADER_ANNUAL_RE = re.compile(r"annual\s+data", re.I)
# "First Eleven Months (…)", "First 11 Months (…)", and the FY 2080/81
# variant "First Eleven(Shrawan-Jestha)" — the word "Months" is optional.
HEADER_MONTHS_RE = re.compile(r"based\s+on\s+first\s+([a-z]+|\d{1,2})\s*(?:months?)?\s*\(", re.I)
HEADER_SINGLE_RE = re.compile(r"first\s+month\s*\(\s*([a-z]+)\s*\)", re.I)
HEADER_RANGE_RE = re.compile(r"\(\s*[a-z]+\s*-\s*([a-z()]+)\s*\)", re.I)


def parse_coverage(header: str) -> tuple[int, int]:
    """(bs_fy_start, month_count 1..12) from the sheet's coverage line."""
    fy = HEADER_FY_RE.search(header)
    if not fy:
        raise ValueError(f"cannot find 'FY YYYY/YY' in header: {header!r}")
    bs_fy_start = int(fy.group(1))
    if HEADER_ANNUAL_RE.search(header):
        return bs_fy_start, 12
    # single-month file first — "First Month (Shrawan)" must not fall into
    # the months-count regex (whose "Months" word is optional)
    if HEADER_SINGLE_RE.search(header):
        return bs_fy_start, 1
    m = HEADER_MONTHS_RE.search(header)
    if m:
        word = m.group(1).lower()
        if word.isdigit():
            count = int(word)  # some files write "First 11 Months"
            if not 1 <= count <= 12:
                raise ValueError(f"month count {count} out of range in header: {header!r}")
        elif word in ORDINAL_MONTH_COUNT:
            count = ORDINAL_MONTH_COUNT[word]
        else:
            raise ValueError(f"unknown month-count word {word!r} in header: {header!r}")
        # cross-check against the "(Shrawan-X)" range when present
        rng = HEADER_RANGE_RE.search(header)
        if rng:
            end = re.sub(r"[^a-z]", "", rng.group(1).lower())
            end_idx = FY_MONTH_BY_NAME.get(end)
            if end_idx is not None and end_idx != count:
                raise ValueError(
                    f"header month-count {count} contradicts range month "
                    f"{end!r} (={end_idx}): {header!r}"
                )
        return bs_fy_start, count
    raise ValueError(f"cannot parse coverage from header: {header!r}")


def parse_workbook(data: bytes, origin: str) -> tuple[int, int, dict[str, dict[str, float]]]:
    """Parse one FTS workbook.

    Returns (bs_fy_start, month_count, {"Whole": {bucket: cum_qty},
    "3-Wheelers": {...}}) — cumulative FYTD quantities.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if "Imports_By_Commodity" in name and "Partner" not in name:
            sheet = wb[name]
            break
    if sheet is None:
        raise ValueError(f"{origin}: no Imports_By_Commodity sheet (sheets: {wb.sheetnames})")

    header_line = None
    cols = None            # {'code': i, 'desc': i, 'unit': i, 'qty': i}
    cum = {v: dict.fromkeys(BUCKETS, 0.0) for v in VARIANT_CONFIG}
    warned: list[str] = []

    for row in sheet.iter_rows(values_only=True):
        if row is None:
            continue
        cells = ["" if c is None else str(c).strip() for c in row]
        if cols is None:
            if header_line is None:
                joined = " ".join(cells)
                if "based on" in joined.lower():
                    header_line = joined
            if cells and cells[0].lower() == "hscode":
                lower = [c.lower() for c in cells]
                cols = {
                    "code": 0,
                    "desc": lower.index("description"),
                    "unit": lower.index("unit"),
                    "qty": lower.index("quantity"),
                }
            continue
        code = cells[cols["code"]]
        if code.endswith(".0"):  # numerically-typed cells stringify as '87038011.0'
            code = code[:-2]
        if not code.startswith("8703") or not code.isdigit() or len(code) != 8:
            continue
        desc = cells[cols["desc"]]
        unit = cells[cols["unit"]]
        raw_qty = row[cols["qty"]]
        try:
            qty = float(raw_qty)
        except (TypeError, ValueError):
            raise ValueError(f"{origin}: non-numeric quantity {raw_qty!r} for {code}")
        if qty < 0:
            raise ValueError(f"{origin}: negative quantity {qty} for {code}")
        if unit.upper() != "PCS":
            warned.append(f"{code} unit={unit!r} (expected PCS)")

        bucket = PREFIX_BUCKET.get(code[:6])
        if bucket is None:
            warned.append(f"unknown sub-code {code} ({desc[:60]!r}) -> OTHERS")
            bucket = "OTHERS"
        # legacy files park electrics under 8703.90 before 8703.80 existed
        if bucket == "OTHERS" and ELECTRIC_RE.search(desc):
            bucket = "BEV"
        variant = "3-Wheelers" if THREE_WHEELER_RE.search(desc) else "Whole"
        cum[variant][bucket] += qty

    if cols is None:
        raise ValueError(f"{origin}: HSCode header row not found")
    if header_line is None:
        raise ValueError(f"{origin}: coverage header line not found")
    for w in warned:
        print(f"    WARNING {origin}: {w}")

    fy, months = parse_coverage(header_line)
    return fy, months, cum


# --------------------------------------------------------------------------
# Assembly: cumulative snapshots -> monthly rows
# --------------------------------------------------------------------------

def monthly_rows(fy: int, snapshots: dict[int, dict[str, dict[str, float]]]):
    """{month_count: cum} -> {variant: {period: {bucket: qty}}}.

    Requires a contiguous 1..K month set. When both an 'upto Ashadh' and an
    annual file exist for month 12, the caller keeps the annual one.
    """
    months = sorted(snapshots)
    if months != list(range(1, len(months) + 1)):
        raise ValueError(f"FY {fy}: non-contiguous month set {months} — cannot diff")
    out = {v: {} for v in VARIANT_CONFIG}
    prev = {v: dict.fromkeys(BUCKETS, 0.0) for v in VARIANT_CONFIG}
    for k in months:
        for variant in VARIANT_CONFIG:
            row = {}
            for b in BUCKETS:
                delta = snapshots[k][variant][b] - prev[variant][b]
                if delta < -1e-6:
                    raise ValueError(
                        f"FY {fy} month {k} {variant} {b}: cumulative value "
                        f"decreased by {-delta:.2f} — upstream revision, "
                        f"needs a human look"
                    )
                row[b] = max(delta, 0.0)
            prev[variant] = snapshots[k][variant]
            out[variant][period_for(fy, k)] = row
    return out


def fmt(x: float) -> str:
    """Format a quantity: integers as 'N.0' (repo style), else 2 decimals."""
    if abs(x - round(x)) < 1e-6:
        return f"{round(x):.1f}"
    return f"{x:.2f}"


def upsert_csv(path: Path, variant: str, per_period: dict[str, dict[str, float]],
               force: bool) -> bool:
    """Merge freshly computed periods into the CSV (keyed by period).

    Existing rows for other periods are preserved; rows for recomputed
    periods are replaced. Returns True when the file content changed.
    """
    rows: dict[str, dict[str, str]] = {}
    if path.exists():
        with path.open(newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                rows[row["period"]] = row

    for period, buckets in per_period.items():
        total = sum(buckets.values())
        if total == 0:
            # a published month with zero passenger-vehicle imports would be
            # extraordinary — refuse to write it silently
            raise ValueError(f"{variant} {period}: all-zero month parsed")
        rows[period] = {
            "period": period,
            "time_interval": "monthly",
            "variant": variant,
            "source": SOURCE,
            "BEV": fmt(buckets["BEV"]),
            "PHEV": "",
            "HEV": fmt(buckets["HEV"]),
            "PETROL": fmt(buckets["PETROL"]),
            "DIESEL": fmt(buckets["DIESEL"]),
            "OTHERS": fmt(buckets["OTHERS"]),
            "TOTAL": fmt(total),
            "notes": "",
        }

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_COLUMNS, lineterminator="\n")
    writer.writeheader()
    for period in sorted(rows):
        writer.writerow({c: rows[period].get(c, "") for c in CSV_COLUMNS})
    new_content = buf.getvalue()

    old_content = path.read_text(encoding="utf-8") if path.exists() else None
    if new_content == old_content and not force:
        return False
    path.write_text(new_content, encoding="utf-8")
    return True


# --------------------------------------------------------------------------
# Drivers
# --------------------------------------------------------------------------

def process_snapshots(fy: int, raw: list[tuple[bytes, str]], force: bool) -> list[str]:
    """Parse workbook blobs of one FY, rebuild its months, upsert CSVs.

    Returns the list of variants whose CSV changed.
    """
    snapshots: dict[int, dict] = {}
    for data, origin in raw:
        got_fy, months, cum = parse_workbook(data, origin)
        if got_fy != fy:
            raise ValueError(f"{origin}: header says FY {got_fy}, expected {fy}")
        if months in snapshots:
            # duplicate coverage (e.g. 'upto Ashadh' + annual): keep the
            # later-listed file — pages list annual first, so keep first seen
            print(f"    note: duplicate month-{months} snapshot ({origin}) ignored")
            continue
        snapshots[months] = cum
        print(f"    parsed {origin}: FY {fy}, months 1..{months}")
    per_variant = monthly_rows(fy, snapshots)
    changed = []
    for variant, csv_path in VARIANT_CONFIG.items():
        if upsert_csv(Path(csv_path), variant, per_variant[variant], force):
            changed.append(variant)
    return changed


def run_online(fy_filter: int | None, backfill: bool, force: bool) -> int:
    session = requests.Session()
    session.headers.update(HTTP_HEADERS)
    categories = discover_fy_categories(session)
    years = sorted(categories, reverse=True)
    if fy_filter:
        years = [y for y in years if y == fy_filter]
        if not years:
            print(f"FY {fy_filter} not found on the site (have: {sorted(categories)})")
            return 1
    elif not backfill:
        years = years[:1]  # newest FY only (steady state)

    all_changed: set[str] = set()
    for fy in years:
        print(f"== FY {fy}/{(fy + 1) % 100:02d}")
        content_url = find_fts_content_page(session, fy, categories[fy])
        if content_url is None:
            msg = f"    no FTS statistics item found for FY {fy} (empty/revenue-only category)"
            if backfill:
                print(msg + " — skipped")
                continue
            if fy_filter:
                print(msg)
                return 1
            # steady state at FY rollover: the new FY's category can exist
            # before its first FTS file — fall back to the previous FY
            print(msg + " — falling back to previous FY")
            prev = [y for y in sorted(categories, reverse=True) if y < fy]
            if not prev:
                return 1
            fy = prev[0]
            content_url = find_fts_content_page(session, fy, categories[fy])
            if content_url is None:
                print(f"    previous FY {fy} has no FTS item either — giving up")
                return 1
            print(f"== FY {fy}/{(fy + 1) % 100:02d} (fallback)")
        workbooks = list_workbooks(session, content_url)
        if not backfill and not fy_filter and not force:
            # cheap steady-state skip: the page lists one workbook per
            # published month; when the CSV already carries that many months
            # of this FY there is nothing new — skip the ~25 MB download.
            fy_periods = {period_for(fy, k) for k in range(1, 13)}
            have = 0
            whole_csv = Path(VARIANT_CONFIG["Whole"])
            if whole_csv.exists():
                with whole_csv.open(newline="", encoding="utf-8") as fh:
                    have = sum(1 for r in csv.DictReader(fh) if r["period"] in fy_periods)
            if have >= len(workbooks):
                print(f"    {have} months of FY {fy} already in CSV, "
                      f"{len(workbooks)} files published — nothing new")
                return 0
        raw = []
        for url, label in workbooks:
            print(f"    downloading {label!r}: {url.rsplit('/', 1)[-1][:70]}")
            r = session.get(url, timeout=180)
            r.raise_for_status()
            raw.append((r.content, url.rsplit("/", 1)[-1]))
        all_changed |= set(process_snapshots(fy, raw, force))

    if all_changed:
        print(f"changed variants: {sorted(all_changed)}")
    else:
        print("no changes")
    return 0


def run_offline(from_dir: Path, force: bool) -> int:
    """Parse every .xlsx in a directory (grouped by the FY in their headers)."""
    by_fy: dict[int, list[tuple[bytes, str]]] = {}
    for p in sorted(from_dir.glob("*.xlsx")):
        data = p.read_bytes()
        fy, months, _ = parse_workbook(data, p.name)
        by_fy.setdefault(fy, []).append((data, p.name))
        print(f"  {p.name}: FY {fy} months 1..{months}")
    changed: set[str] = set()
    for fy, raw in sorted(by_fy.items()):
        print(f"== FY {fy} ({len(raw)} files)")
        changed |= set(process_snapshots(fy, raw, force))
    print(f"changed variants: {sorted(changed)}" if changed else "no changes")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--fy", type=int, help="process one specific FY (BS start year, e.g. 2081)")
    ap.add_argument("--backfill", action="store_true",
                    help="process every FY discoverable on the site")
    ap.add_argument("--from-dir", type=Path,
                    help="offline mode: parse local .xlsx files from this directory")
    ap.add_argument("--force", action="store_true",
                    help="rewrite CSVs even when content is unchanged")
    args = ap.parse_args()

    if args.from_dir:
        return run_offline(args.from_dir, args.force)
    return run_online(args.fy, args.backfill, args.force)


if __name__ == "__main__":
    sys.exit(main())
