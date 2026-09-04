#!/usr/bin/env python3
"""
Fetch France passenger-car registration data from the SDES (Service des
données et études statistiques, Ministère de la Transition écologique)
"Immatriculations mensuelles de voitures neuves par motorisation" series and
upsert `data/France.csv` (variant `Whole`).

Why SDES (and not ACEA)
-----------------------
France was the last large market still served by the ACEA aggregate. SDES is
the national registry-statistics branch of the same SIV registry ACEA/PFA/AAA
sit on, published monthly, free and open (Licence Ouverte). The detailed
"motorisation" series is the only SDES product that keeps a real **HEV**
column (the coarse open-data "source d'énergie" taxonomy folds HEV into
petrol/diesel — see docs/architecture/36-source-france.md). It also carries the
**full fuel split back to 2011-01**, where the old ACEA series was sparse /
quarterly-interpolated before ~2025.

Source workbook
---------------
`https://www.statistiques.developpement-durable.gouv.fr/media/<id>/download`
One .xlsx, one sheet, a monthly time series from 2011_01. Columns (row 3):

    Gazole (thermique) | Essence (thermique) |
    hybride gazole non rechargeable | hybride essence non rechargeable |
    gazole (y compris hybrides non rechargeables) |       <- aggregate, IGNORED
    essence (y compris hybrides non rechargeables) |       <- aggregate, IGNORED
    hybride rechargeable | Electrique | Gaz & ND | Total

Footer row is `Source : SDES-RSVERO`. NB the sheet TAB is misnamed one month
behind (e.g. tab `2026_05` for a file whose last row is `2026_06`), so we read
the latest month from the period column, never the tab name.

Mapping to the canonical CSV columns
------------------------------------
    BEV     <- Electrique
    PHEV    <- hybride rechargeable
    HEV     <- hybride gazole non rechargeable + hybride essence non rechargeable
    PETROL  <- Essence (thermique)          (pure petrol, excl. non-recharg. hybrids)
    DIESEL  <- Gazole (thermique)           (pure diesel,  excl. non-recharg. hybrids)
    OTHERS  <- Gaz & ND                      (LPG/CNG + non-déterminé)
    TOTAL   <- Total   (asserted == sum of the seven base columns above)

The two "(y compris hybrides)" columns are convenience aggregates
(= thermique + non-rechargeable hybride) and are deliberately NOT summed in.

Consistency with the ACEA target (documented, not tuned away)
-------------------------------------------------------------
Totals reconcile with ACEA/AAA to within ~0.2 % (the SDES "hors provisoires et
transit temporaire" residual). Per-fuel there is a small, systematic
definitional step: SDES puts ~3 % less in HEV / ~3.5 % more in PETROL than
ACEA (ACEA/PFA count 48 V mild-hybrids as HEV; SDES counts them as pure
petrol — SDES is actually closer to this project's MHEV->ICE glossary
convention), and SDES OTHERS is higher because "Gaz & ND" carries the
not-yet-coded "ND" bucket. BEV share is essentially unchanged (~+0.3 pp).
See docs/architecture/36-source-france.md § "Fachlich".

Usage
-----
    python scripts/fetch_france.py --file <local.xlsx>   # parse a local workbook
    python scripts/fetch_france.py --url  <media url>    # download one workbook
    python scripts/fetch_france.py                       # resolve latest, then fetch
    #   optional: --out data/France.csv (default) · --dry-run · --page <hub url>

The live download resolves the current `media/<id>/download` by probing the SDES
hub and the per-year "Données <YYYY>" pages (the id rotates every publication)
and scoring each média link by its label. That resolution is the one part that
cannot be tested from the Claude sandbox (egress to every French host is
policy-denied) — it is validated from GitHub Actions, and on failure the error
dumps a per-page média/marker diagnostic so the next fix is one-shot. The parser
and the upsert are fully validated locally against the published workbook.
Invoked by `.github/workflows/fetch-france.yml`.
"""
import argparse
import csv
import datetime
import io
import re
import sys
from html import unescape
from pathlib import Path

import requests
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
OUT_DEFAULT = REPO / "data" / "France.csv"
SOURCE = "SDES"

# Stable landing page that links the current série workbook. The média id in the
# download URL rotates each publication, so we resolve it from here rather than
# hardcode it. VALIDATE IN GHA (see module docstring).
LANDING_PAGE = (
    "https://www.statistiques.developpement-durable.gouv.fr/"
    "immatriculation-des-vehicules-routiers"
)
# SDES média download URLs are opaque — "/media/<id>/download" carries NO
# filename — so the série is identified by the human-readable LABEL near the
# reference, not by the URL, and the média id rotates every publication. We
# match "/media/<id>/download" in ANY context (tolerating JSON "\/" escaping),
# so an id sitting in an inline JSON/state blob is found as well as a plain
# href. The downloads do not live on the topic hub (a nav page with no
# attachments) but on the per-year "Données <YYYY> sur les immatriculations des
# véhicules" page, which resolve_latest_url derives from the hub and probes too.
MEDIA_RE = re.compile(r"media[\\/]+(\d+)[\\/]+download", re.I)
_YEAR_PAGE = "/donnees-{year}-sur-les-immatriculations-des-vehicules"
_VP_POS = ("particuli",)                         # voitures particulières (VP)
_ENERGY_POS = ("énergie", "energie", "motorisation")
# On the monthly VP publication the long-series workbook is labelled generically
# as "…séries mensuelles des immatriculations" — the product name ("VP neuves par
# énergie") lives only in the downloaded filename — so match that phrase, with the
# VP context supplied by the page URL (voitures-particulieres-…).
_SERIES_POS = ("séries mensuelles", "series mensuelles",
               "série mensuelle", "serie mensuelle")
# Anything that is NOT the monthly VP-énergie série: other categories/markets, the
# annual "séries longues … par genre", the complementary/CO2/méthodologie files,
# and the departmental/regional/marque breakdowns.
_REJECT = ("utilitaire", "poids lourd", "autobus", "autocar", "deux-roues",
           "deux roues", "cyclomoteur", "camion", "occasion",
           "méthodolog", "methodolog", "longues", "annuel", "genre",
           "complémentaire", "complementaire", "co2", "bonus", "malus",
           "marque", "région", "region", "départ", "depart", "commune")
# Markers that reveal a page's real data structure when no usable média id is
# found — dumped in the failure diagnostic so the next fix is informed.
_MARKERS = ("/media/", "download", "télécharg", "telecharg", ".xlsx", "dido",
            "data.gouv", "application/json", "__nuxt", "__next", "window.__",
            "data-drupal", "drupal-media")
_SCRIPT_SRC = re.compile(r'<script[^>]+src="([^"]+)"', re.I)

CSV_COLUMNS = [
    "period", "time_interval", "variant", "source",
    "BEV", "PHEV", "HEV", "PETROL", "DIESEL", "OTHERS", "TOTAL", "notes",
]
PERIOD_RE = re.compile(r"^\s*(\d{4})_(\d{2})\s*$")


def _norm(s) -> str:
    """Normalise a header cell: lowercase, collapse whitespace/newlines."""
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _column_index(header: list, *, contains=(), equals=None, exclude=()) -> int:
    """Find the single column whose normalised header matches. Raises if 0 or >1."""
    hits = []
    for i, cell in enumerate(header):
        h = _norm(cell)
        if not h:
            continue
        if equals is not None and h != equals:
            continue
        if any(tok not in h for tok in contains):
            continue
        if any(tok in h for tok in exclude):
            continue
        hits.append(i)
    if len(hits) != 1:
        raise ValueError(
            f"expected exactly one column for "
            f"{equals or contains!r} (excl {exclude!r}); found {hits} "
            f"in header {[_norm(c) for c in header]}"
        )
    return hits[0]


def _resolve_columns(header: list) -> dict:
    """Map canonical fuel keys -> source column indices, by header text."""
    return {
        "DIESEL": _column_index(header, contains=("gazole", "thermique")),
        "PETROL": _column_index(header, contains=("essence", "thermique")),
        # the two non-rechargeable hybrid columns are summed into HEV; the
        # "(y compris hybrides ...)" aggregates must be excluded — they also
        # contain "hybride"/"gazole"/"non recharg".
        "HEV_D": _column_index(header, contains=("hybride", "gazole", "non recharg"),
                               exclude=("y compris",)),
        "HEV_E": _column_index(header, contains=("hybride", "essence", "non recharg"),
                               exclude=("y compris",)),
        "PHEV": _column_index(header, equals="hybride rechargeable"),
        "BEV": _column_index(header, contains=("electrique",)),
        "OTHERS": _column_index(header, contains=("gaz", "nd"), exclude=("gazole",)),
        "TOTAL": _column_index(header, equals="total"),
    }


def _find_header_and_period_col(rows: list) -> tuple:
    """Locate the header row (the one carrying 'Total') and the period column."""
    for r_idx, row in enumerate(rows):
        if any(_norm(c) == "total" for c in row):
            # period column: the one whose *next* data rows look like YYYY_MM
            for c_idx, _ in enumerate(row):
                for probe in rows[r_idx + 1: r_idx + 6]:
                    if c_idx < len(probe) and PERIOD_RE.match(str(probe[c_idx] or "")):
                        return r_idx, list(row), c_idx
    raise ValueError("could not locate the header row (no 'Total' column found)")


def _num(x) -> float:
    return 0.0 if x in (None, "") else float(x)


def _rows_to_records(rows: list, h_idx: int, period_col: int, cols: dict) -> list:
    """Turn the data rows below the header into canonical record dicts."""
    out = []
    for row in rows[h_idx + 1:]:
        m = PERIOD_RE.match(str(row[period_col] if period_col < len(row) else "" or ""))
        if not m:
            continue  # blank separators and the 'Source : SDES-RSVERO' footer
        year, month = m.group(1), m.group(2)
        bev = _num(row[cols["BEV"]])
        phev = _num(row[cols["PHEV"]])
        hev = _num(row[cols["HEV_D"]]) + _num(row[cols["HEV_E"]])
        petrol = _num(row[cols["PETROL"]])
        diesel = _num(row[cols["DIESEL"]])
        others = _num(row[cols["OTHERS"]])
        total = _num(row[cols["TOTAL"]])
        base = bev + phev + hev + petrol + diesel + others
        if abs(base - total) > 1.0:  # the file's own Total must close to the parts
            raise ValueError(
                f"{year}-{month}: parts {base:.0f} != Total {total:.0f} "
                f"(check column mapping)"
            )
        out.append({
            "period": f"{year}-{month}",
            "time_interval": "monthly",
            "variant": "Whole",
            "source": SOURCE,
            "BEV": bev, "PHEV": phev, "HEV": hev,
            "PETROL": petrol, "DIESEL": diesel, "OTHERS": others,
            "TOTAL": total,
            "notes": "",
        })
    return out


def _sheet_preview(rows: list, n: int = 5) -> str:
    """A compact preview of a sheet's first rows for the failure diagnostic."""
    out = []
    for r in rows[:n]:
        cells = [str(c)[:20] for c in (r or ()) if c not in (None, "")]
        out.append(" | ".join(cells[:12]) if cells else "(empty)")
    return " ;; ".join(out)


def parse_workbook(data: bytes) -> list:
    """Parse the SDES workbook bytes -> list of canonical row dicts (period asc).

    The workbook may carry several sheets (one per vehicle category, a cover
    sheet, …), so scan every sheet and use the first that exposes the motorisation
    header (a 'Total' column plus the fuel columns). On total failure, dump each
    sheet's name and first rows so the real layout is visible from one CI run.
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    problems = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        try:
            h_idx, header, period_col = _find_header_and_period_col(rows)
            cols = _resolve_columns(header)
        except ValueError as exc:
            problems.append(f"  sheet {ws.title!r} ({len(rows)} rows): {exc}\n"
                            f"    head: {_sheet_preview(rows)}")
            continue
        out = _rows_to_records(rows, h_idx, period_col, cols)
        if out:
            if len(wb.worksheets) > 1:
                print(f"[france] parsed sheet {ws.title!r} "
                      f"(of {len(wb.worksheets)} in the workbook)")
            return out
        problems.append(f"  sheet {ws.title!r}: header found but no data rows\n"
                        f"    head: {_sheet_preview(rows)}")
    raise ValueError("no sheet matched the SDES motorisation série layout "
                     "(Total + fuel columns):\n" + "\n".join(problems))


def _clean(fragment: str) -> str:
    """Strip HTML tags + entities from a fragment; collapse whitespace."""
    return re.sub(r"\s+", " ", unescape(re.sub(r"<[^>]+>", " ", fragment))).strip()


def _label_for(doc: str, a_start: int, a_end: int) -> str:
    """Descriptive text of a /media/<id>/download reference.

    The id sits INSIDE ``href="/media/<id>/download"``, so the anchor's own tag is
    not the label. On these Drupal pages each download's title is the text that
    precedes its button (a heading or field label) and does not run past the
    previous download link. So bound the window to the gap since the previous
    /download (capped ~500 chars), start it just after a tag boundary so no
    partial tag leaks in, and extend the right edge through this anchor's own
    inner text (to </a>) to catch a "… par énergie" that lives in the link text.
    """
    floor = max(0, a_start - 500)
    prev = doc.rfind("/download", floor, a_start)
    if prev != -1:                           # start past the PREVIOUS download's
        pclose = doc.find("</a>", prev, a_start)   # own title, so it can't bleed
        lo = max(floor, pclose + 4 if pclose != -1 else prev + len("/download"))
    else:
        lo = floor
    gt = doc.find(">", lo, a_start)          # begin after a full tag, not mid-tag
    if gt != -1:
        lo = gt + 1
    close = doc.find("</a>", a_end)          # extend through this link's own text
    hi = close if (close != -1 and close - a_end < 220) else a_end + 100
    return _clean(doc[lo:hi])


# French month names as they appear in the SDES slug (accents stripped by Drupal
# pathauto: février -> fevrier, août -> aout, décembre -> decembre).
_MONTHS_FR = ("janvier", "fevrier", "mars", "avril", "mai", "juin", "juillet",
              "aout", "septembre", "octobre", "novembre", "decembre")
# The national VP-neuves-par-énergie "série" workbook is the "données" of the
# monthly "Motorisations des véhicules légers neufs" publication (NOT the
# "immatriculations-de-voitures-particulieres-neuves-en-…" page, which carries
# only the regional/annual/methodology files — see docs/architecture/36).
_MOTOR_PAGE = ("/motorisations-des-vehicules-legers-neufs-emissions-de-co2-"
               "et-bonus-ecologique-{month}-{year}")


def _statinfo_slugs(n: int = 6) -> list:
    """Recent monthly "Motorisations" publication slugs, newest first.

    The série workbook (2011 → latest month, one file) is re-attached to each
    monthly Motorisations publication with a fresh média id, so the newest page
    carries the current série. This publication lags the headline immatriculations
    note, so walk back n months; probe the base slug and the pathauto "-0" dedup
    variant. Start one month back (the current month is not published yet)."""
    today = datetime.date.today()
    y, m = today.year, today.month
    slugs = []
    for _ in range(n):
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        base = _MOTOR_PAGE.format(month=_MONTHS_FR[m - 1], year=y)
        slugs += [base, base + "-0"]
    return slugs


def _candidate_pages(page: str) -> list:
    """Pages to probe for the current série média link, in priority order.

    1. the recent monthly VP StatInfo pages — where the continuous série is
       re-attached each month (the happy path; probing stops at the first strong
       match, so normally only the newest existing month is fetched);
    2. the topic hub as given (a nav page, usually no attachments — cheap safety);
    3. the per-year "Données <YYYY>" pages (annual tables; a last-resort fallback,
       current + previous year to survive the New-Year rollover).
    """
    year = datetime.date.today().year
    seen, pages = set(), []
    for path in _statinfo_slugs():
        pages.append(requests.compat.urljoin(page, path))
    pages.append(page)
    for y in (year, year - 1):
        pages.append(requests.compat.urljoin(page, _YEAR_PAGE.format(year=y)))
    return [p for p in pages if not (p in seen or seen.add(p))]


def _page_vp(url: str) -> bool:
    """The page is a voitures-particulières publication — it supplies the VP
    context for a generically-labelled "séries mensuelles" download."""
    return "voitures-particulieres" in url.lower()


def _score_label(label: str, vp_ctx: bool = False) -> int:
    low = label.lower()
    vp = vp_ctx or any(k in low for k in _VP_POS)
    series = any(k in low for k in _SERIES_POS)
    energy = any(k in low for k in _ENERGY_POS)
    return (2 * vp + 3 * series + 2 * energy
            + ("immatriculation" in low)
            - 6 * any(k in low for k in _REJECT))


def _eligible(label: str, vp_ctx: bool = False) -> bool:
    """The monthly VP-énergie série: passenger-car context (from the label or the
    page URL), a monthly-series or énergie signal, and none of the reject markers
    (other categories, the annual/genre séries, méthodologie, CO2, the regional/
    marque breakdowns)."""
    low = label.lower()
    vp = vp_ctx or any(k in low for k in _VP_POS)
    signal = (any(k in low for k in _SERIES_POS)
              or any(k in low for k in _ENERGY_POS))
    return vp and signal and not any(k in low for k in _REJECT)


def _scan_page(session: requests.Session, url: str) -> dict:
    """Fetch one page and collect every distinct /media/<id>/download it references.

    Returns a diagnostics dict: byte size, média candidates (id -> best
    (score, label, eligible)), marker counts and the first script srcs — enough to
    tell, from a single CI run, whether the série link is present in the HTML or
    the page is a JS shell that pulls it from elsewhere.
    """
    try:
        r = session.get(url, timeout=60)
        r.raise_for_status()
    except Exception as exc:  # noqa: BLE001 - record and try the next candidate
        return {"url": url, "error": str(exc)}
    doc = r.text
    low = doc.lower()
    vp_ctx = _page_vp(url)
    best: dict = {}
    for m in MEDIA_RE.finditer(doc):
        mid = int(m.group(1))
        label = _label_for(doc, m.start(), m.end())
        raw = re.sub(r"\s+", " ", doc[max(0, m.start() - 180): m.end() + 120])
        cand = (_score_label(label, vp_ctx), label, _eligible(label, vp_ctx), raw)
        if mid not in best or cand[0] > best[mid][0]:
            best[mid] = cand
    return {
        "url": url, "bytes": len(doc), "media": best,
        "markers": {k: low.count(k) for k in _MARKERS},
        "scripts": _SCRIPT_SRC.findall(doc)[:6],
    }


def resolve_latest_url(session: requests.Session, page: str) -> str:
    """Resolve the current 'VP neuves par énergie' série média URL.

    Probes the recent monthly VP StatInfo pages (then the hub and per-year pages);
    among every /media/<id>/download found, picks the one whose surrounding label
    reads as the monthly VP série — "…séries mensuelles des immatriculations" on a
    voitures-particulières page, or an explicit "…par énergie" — while rejecting
    the annual/genre, méthodologie, complémentaire and regional/marque files.
    Newest id wins ties; a strong match stops the probing early. On failure the
    error dumps, per page, what was found (média ids + labels + raw markup) and the
    page's structural markers, so the next fix is one-shot, not a blind guess.
    """
    scans, winner = [], None  # winner: (eligible, score, media_id, url, label)
    for u in _candidate_pages(page):
        sc = _scan_page(session, u)
        scans.append(sc)
        for mid, (score, label, elig, _raw) in sc.get("media", {}).items():
            url = requests.compat.urljoin(sc["url"], f"/media/{mid}/download")
            cand = (elig, score, mid, url, label)
            if winner is None or cand[:3] > winner[:3]:
                winner = cand
        if winner and winner[0] and winner[1] >= 5:  # strong match — stop probing
            break
    if winner and winner[0] and winner[1] >= 3:
        print(f"[france] resolved {winner[3]} (score {winner[1]}): {winner[4][:120]}")
        return winner[3]

    # ---- nothing usable: dump everything one CI run needs to see ----
    lines = []
    for sc in scans:
        if "error" in sc:
            lines.append(f"  {sc['url']} -> FETCH ERROR: {sc['error']}")
            continue
        mk = " ".join(f"{k}={v}" for k, v in sc["markers"].items() if v)
        lines.append(f"  {sc['url']} [{sc['bytes']}B, {len(sc['media'])} média id(s)]  "
                     f"markers: {mk or 'none'}")
        for mid, (score, label, elig, raw) in sorted(sc["media"].items(),
                                                      key=lambda kv: -kv[1][0])[:10]:
            lines.append(f"      [{score}{'*' if elig else ' '}] media/{mid}: {label[:90]}")
            lines.append(f"          raw: …{raw[:200]}…")
        if not sc["media"] and sc["scripts"]:
            lines.append(f"      scripts: {', '.join(s[:70] for s in sc['scripts'])}")
    raise ValueError("could not identify the VP-énergie série média link.\n"
                     + "\n".join(lines)
                     + "\nWorkaround: pass --url with the média URL explicitly.")


def fetch_bytes(args, session: requests.Session) -> bytes:
    if args.file:
        return Path(args.file).read_bytes()
    url = args.url or resolve_latest_url(session, args.page)
    print(f"[france] downloading {url}")
    r = session.get(url, timeout=120)
    r.raise_for_status()
    fetch_bytes.url = url  # remember for the notes column
    return r.content


def diagnose(session: requests.Session, page: str) -> int:
    """Download every média on the newest StatInfo page and report its identity.

    Which of the page's downloads is the national VP-énergie série cannot be told
    from the link label alone (the labels are generic and rotate ids), so pull each
    file and print its Content-Disposition filename plus its sheet names — the file
    named "…serie_vp_neuves_par_ener…" with a single month-tab sheet is the target.
    Run via the workflow's `diagnose` input; it never writes the CSV.
    """
    for u in _candidate_pages(page):
        sc = _scan_page(session, u)
        media = sc.get("media", {})
        if not media:
            continue
        print(f"[diag] {u} — {len(media)} média")
        for mid in sorted(media):
            label = media[mid][1]
            url = requests.compat.urljoin(u, f"/media/{mid}/download")
            try:
                r = session.get(url, timeout=120)
                r.raise_for_status()
            except Exception as exc:  # noqa: BLE001
                print(f"[diag]   media/{mid}: DOWNLOAD ERROR {exc}")
                continue
            cd = r.headers.get("Content-Disposition", "")
            line = (f"[diag]   media/{mid} ({len(r.content)}B) label={label[:45]!r} "
                    f"cd={cd[:120]!r}")
            try:
                wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
                first = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))[:4]
                line += f"\n[diag]       sheets={wb.sheetnames[:15]} head0={_sheet_preview(first)}"
            except Exception as exc:  # noqa: BLE001 - not every média is an xlsx
                line += f"  [not an xlsx: {exc}]"
            print(line)
        return 0
    print("[diag] no candidate page exposed any média link")
    return 1


def upsert(out_path: Path, new_rows: list, *, dry_run: bool, src_url: str) -> None:
    """Write new_rows into out_path, honouring the ACEA-courtesy rule.

    Overwrite an existing row only if its source is exactly ``ACEA`` (the
    fallback we are replacing) or already ``SDES``. Never touch a row written by
    any other source. New months are appended.
    """
    existing: dict[str, dict] = {}
    order: list[str] = []
    if out_path.exists():
        with out_path.open(encoding="utf-8", newline="") as fh:
            for row in csv.DictReader(fh):
                existing[row["period"]] = row
                order.append(row["period"])

    latest = max(r["period"] for r in new_rows)
    changed = skipped = added = 0
    for nr in new_rows:
        p = nr["period"]
        # only the newest row carries the source URL in notes (repo convention)
        nr = {**nr, "notes": (src_url if (p == latest and src_url) else "")}
        old = existing.get(p)
        if old is None:
            existing[p] = nr
            order.append(p)
            added += 1
        elif (old.get("source") or "").strip() in (SOURCE, "ACEA"):
            # full overwrite: the row is now SDES-sourced, so a stale ACEA note
            # (e.g. an acea.auto PDF URL) must not survive on it.
            existing[p] = nr
            changed += 1
        else:
            skipped += 1

    order = sorted(dict.fromkeys(order))
    print(f"[france] {added} added, {changed} updated, {skipped} preserved "
          f"(non-ACEA/SDES source) -> {out_path}")
    if dry_run:
        print("[france] --dry-run: not writing")
        return
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=CSV_COLUMNS)
        w.writeheader()
        for p in order:
            w.writerow({c: existing[p].get(c, "") for c in CSV_COLUMNS})


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Fetch France (SDES) registrations.")
    ap.add_argument("--file", help="parse a local .xlsx instead of downloading")
    ap.add_argument("--url", help="download this exact média .xlsx URL")
    ap.add_argument("--page", default=LANDING_PAGE, help="landing page to resolve the média link")
    ap.add_argument("--out", default=str(OUT_DEFAULT), help="output CSV (default data/France.csv)")
    ap.add_argument("--dry-run", action="store_true", help="report changes, do not write")
    ap.add_argument("--diagnose", action="store_true",
                    help="download every média on the StatInfo page and report each "
                         "file's filename + sheets (identify the série link); no write")
    args = ap.parse_args(argv)

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (LeRaffl-Gallery fetch_france)"})
    fetch_bytes.url = ""
    if args.diagnose:
        try:
            return diagnose(session, args.page)
        except Exception as exc:  # noqa: BLE001
            print(f"[france] ERROR: {exc}", file=sys.stderr)
            return 1
    try:
        data = fetch_bytes(args, session)
        rows = parse_workbook(data)
    except Exception as exc:  # noqa: BLE001 - surface a clean CI failure
        print(f"[france] ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"[france] parsed {len(rows)} months: {rows[0]['period']} .. {rows[-1]['period']}")
    upsert(Path(args.out), rows, dry_run=args.dry_run, src_url=getattr(fetch_bytes, "url", "") or (args.url or ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
