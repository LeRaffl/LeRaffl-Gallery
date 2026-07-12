#!/usr/bin/env python3
"""
Morocco (AIVAM) fetcher — currently PROBE ONLY.

Usage
-----
    python scripts/fetch_morocco.py --probe
        Run from CI (the dev sandbox has no external network). Crawls
        AIVAM's publication pages, lists downloadable PDFs, downloads the
        most recent one(s) and dumps their text — to answer the open
        question from docs/architecture/29-expansion-candidates.md: does
        AIVAM's own monthly PDF carry the per-energy split (BEV / PHEV /
        HEV / MHEV / petrol / diesel) that the Moroccan press quotes, or
        does that table only exist in press relays?

AIVAM (Association des Importateurs de Véhicules au Maroc) is the importers'
association; Morocco's formal new-vehicle market flows through its members,
including the Chinese brands, so completeness is expected to be OK.
"""
import argparse
import io
import re

import requests

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36",
      "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.7"}

# Probe v1 (2026-07-12): docs page lists a 1-page monthly press PDF (brand
# table only, NO energy split), the annual "Bilan" deck (has a "Focus NEV"
# section), a monthly XLS ("Statistiques des ventes <Month> FP.xls") and a
# separate stats portal at statistique.aivam.ma. Probe v2 opens the XLS and
# the portal.
PAGES = [
    "https://www.aivam.ma/fr/documentation-et-etudes",
]
XLS_URL = "https://www.aivam.ma/sites/default/files/2025-11/Statistiques%20des%20ventes%20Octobre%202025%20FP.xls"
PORTAL = "https://statistique.aivam.ma"


def probe() -> None:
    print("=== PROBE: AIVAM (Morocco) publications ===", flush=True)
    s = requests.Session()
    s.headers.update(UA)

    pdf_links: list[str] = []
    for url in PAGES:
        try:
            r = s.get(url, timeout=60)
            print(f"\n[{url}] HTTP {r.status_code} final={r.url} len={len(r.text)}", flush=True)
            if not r.ok:
                continue
            links = re.findall(r'href="([^"]+)"', r.text)
            pdfs = [l for l in links if ".pdf" in l.lower()]
            others = [l for l in links if any(k in l.lower() for k in ("statist", "document", "etude", "marche", "vente"))
                      and ".pdf" not in l.lower()]
            print(f"  pdf links ({len(pdfs)}):")
            for l in pdfs[:25]:
                print(f"    {l}")
            print(f"  stats-ish page links ({len(others)}):")
            for l in sorted(set(others))[:15]:
                print(f"    {l}")
            for l in pdfs:
                full = l if l.startswith("http") else ("https://www.aivam.ma" + (l if l.startswith("/") else "/" + l))
                if full not in pdf_links:
                    pdf_links.append(full)
        except requests.RequestException as e:
            print(f"\n[{url}] FAILED: {e}", flush=True)

    # --- The monthly XLS (the likely fetcher target)
    print(f"\n=== XLS: {XLS_URL}", flush=True)
    try:
        r = s.get(XLS_URL, timeout=90)
        print(f"  HTTP {r.status_code}, {len(r.content):,} bytes, "
              f"content-type={r.headers.get('content-type')}")
        if r.ok:
            head = r.content[:8]
            print(f"  magic: {head!r}")
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=r.content)
                for name in wb.sheet_names():
                    sh = wb.sheet_by_name(name)
                    print(f"  --- sheet {name!r}: {sh.nrows} rows x {sh.ncols} cols")
                    for i in range(min(15, sh.nrows)):
                        print(f"    {[sh.cell_value(i, j) for j in range(min(12, sh.ncols))]}")
            except Exception as e:
                print(f"  xlrd failed ({e}); trying openpyxl")
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
                    for name in wb.sheetnames:
                        ws = wb[name]
                        print(f"  --- sheet {name!r}: {ws.max_row} rows x {ws.max_column} cols")
                        for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
                            print(f"    {list(row)[:12]}")
                except Exception as e2:
                    print(f"  openpyxl failed too: {e2}")
    except requests.RequestException as e:
        print(f"  FAILED: {e}")

    # --- The stats portal: Angular SPA (probe v2) — mine the JS bundle for
    # its REST endpoints, then try them.
    print(f"\n=== PORTAL: {PORTAL}", flush=True)
    try:
        r = s.get(PORTAL, timeout=60)
        print(f"  HTTP {r.status_code} final={r.url}")
        bundles = re.findall(r'src="((?:main|runtime|scripts)[^"]+\.js)"', r.text)
        endpoints: set[str] = set()
        for b in bundles:
            burl = f"{PORTAL}/{b}"
            jr = s.get(burl, timeout=60)
            print(f"  bundle {b}: HTTP {jr.status_code}, {len(jr.text):,} chars")
            if not jr.ok:
                continue
            js = jr.text
            endpoints.update(re.findall(r'https?://[A-Za-z0-9._/-]*aivam[A-Za-z0-9._/-]*', js))
            endpoints.update(re.findall(r'"(/?api/[A-Za-z0-9._/-]{2,60})"', js))
            endpoints.update(re.findall(r"'(/?api/[A-Za-z0-9._/-]{2,60})'", js))
            endpoints.update(re.findall(r'"([A-Za-z0-9._-]*(?:energie|marque|vente|statisti|segment)[A-Za-z0-9._/-]*)"', js, re.I))
        print(f"  extracted endpoint candidates ({len(endpoints)}):")
        for e in sorted(endpoints)[:60]:
            print(f"    {e}")
        # Try the most promising absolute URLs / api paths
        tried = 0
        for e in sorted(endpoints):
            if tried >= 10:
                break
            url = e if e.startswith("http") else f"{PORTAL}/{e.lstrip('/')}"
            if not any(k in url.lower() for k in ("api", "energie", "vente", "statisti")):
                continue
            tried += 1
            try:
                ar = s.get(url, timeout=45)
                print(f"  GET {url} -> {ar.status_code} "
                      f"({ar.headers.get('content-type')}) head={ar.text[:200].replace(chr(10), ' ')}")
            except requests.RequestException as ex:
                print(f"  GET {url} -> FAILED {ex}")
    except requests.RequestException as e:
        print(f"  FAILED: {e}")

    print("\n=== PROBE DONE ===")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--probe", action="store_true")
    args = ap.parse_args()
    if args.probe:
        probe()
        return
    raise SystemExit("Only --probe is implemented until the source shape is confirmed.")


if __name__ == "__main__":
    main()
