#!/usr/bin/env python3
"""
Fetch Poland new registration data from PZPM's public eRegistrations workbook
and upsert per-variant CSVs.

Usage
-----
    python scripts/fetch_poland.py [--variant {whole,vans,hdv,buses,all}] [--force]
    python scripts/fetch_poland.py --xlsx PATH --period YYYY-MM [--force]   # parse a local file

Output files (one per PZPM vehicle category; all parsed from the "Ogółem" sheet)
--------------------------------------------------------------------------------
    data/Poland.csv         <- Whole  OSOBOWE                       (passenger cars, M1)
    data/Poland_Vans.csv    <- Vans   SAMOCHODY DOSTAWCZE           (LCV <=3.5t, N1)
    data/Poland_HDV.csv     <- HDV    SAMOCHODY CIĘŻAROWE POW. 3,5T (trucks >3.5t, N2/N3)
    data/Poland_Buses.csv   <- Buses  AUTOBUSY                      (buses, M2/M3)

Source
------
PZPM (Polski Związek Przemysłu Motoryzacyjnego) publishes a monthly
eRegistrations workbook on https://www.pzpm.org.pl/en/Electromobility/eRegistrations
around the 7th of the following month, based on the Central Register of Vehicles
(CEP). The page links a single XLSX ("PZPM_eRejestracje - tabele MM.YYYY.xlsx")
whose /content/download/<id>/<id>/file/ IDs change every month, so the URL must
be discovered by scraping the page (there is no stable URL and no API).

Only the workbook's "Ogółem" (Overall) sheet is reliably updated each month — the
other sheets (brand/model rankings, "Paliwa_...") are stale 2023 template tabs
and must NOT be parsed. "Ogółem" gives, per vehicle category, the current month's
count by drive type plus a year-to-date column (the YTD column is ignored here;
only the current month is taken).

The workbook holds ONLY the current month (no per-month history). The Whole
(passenger) history back to 2010 already lives in data/Poland.csv from the ACEA
pipeline; PZPM is the upstream CEP source behind those ACEA numbers (verified:
PZPM OSOBOWE Apr-2026 = ACEA Poland Apr-2026, to the unit). Going forward PZPM
owns the Whole row (source := "PZPM"); the commercial variants start thin
(current month onward) and accumulate over time, mirroring Portugal.

Drive-type -> canonical column (per "Ogółem" row label, ASCII-folded match):
    Benzyna           -> PETROL
    Diesel            -> DIESEL
    Elektryczne       -> BEV
    Hybrydowe plug-in -> PHEV
    Hybrydowe         -> HEV           (exact; full/mild hybrids)
    OTHERS            -> TOTAL - (BEV+PHEV+HEV+PETROL+DIESEL)
                         (residual captures LPG, Wodorowe/FCEV, CNG/LNG, and — for
                          the commercial variants, where PZPM reports a single
                          combined "Hybrydowe / hybrydowe plug-in" figure that
                          can't be split — the hybrids too)

Columns a category does not report separately are written empty ("" = not
reported), not 0 — e.g. Vans/HDV have no PHEV/HEV split, HDV has no PETROL.

Cross-check: ACEA leaves Poland to this workflow (Poland is no longer in
scripts/fetch_acea.py's country list). See docs/architecture/22-source-poland.md.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import urllib.parse
from io import BytesIO
from pathlib import Path

import openpyxl
import requests

PAGE_URL = "https://www.pzpm.org.pl/en/Electromobility/eRegistrations"
HOST = "https://www.pzpm.org.pl"
SOURCE = "PZPM"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"

# Variant -> ("Ogółem" category header, output CSV). The header is matched after
# ASCII-folding + upper-casing, so diacritics (Ę, Ż, Ó, ...) don't matter.
VARIANT_CONFIG = {
    "Whole": {"header": "OSOBOWE",                       "csv": "data/Poland.csv"},
    "Vans":  {"header": "SAMOCHODY DOSTAWCZE",           "csv": "data/Poland_Vans.csv"},
    "HDV":   {"header": "SAMOCHODY CIEZAROWE POW. 3,5T", "csv": "data/Poland_HDV.csv"},
    "Buses": {"header": "AUTOBUSY",                      "csv": "data/Poland_Buses.csv"},
}
# Category headers in "Ogółem" we recognise but intentionally skip (so their
# drive-type sub-rows are not misattributed to the preceding variant).
SKIP_HEADERS = {"SAMOCHODY CIEZAROWE OD 12T", "MOTOCYKLE", "MOTOROWERY"}

# Drive-type row label (ASCII-folded, lower-cased, exact) -> canonical column.
FUEL_MAP = {
    "benzyna": "PETROL",
    "diesel": "DIESEL",
    "elektryczne": "BEV",
    "hybrydowe plug-in": "PHEV",
    "hybrydowe": "HEV",
}

CSV_COLUMNS = [
    "period", "time_interval", "variant", "source",
    "BEV", "PHEV", "HEV", "PETROL", "DIESEL", "OTHERS", "TOTAL", "notes",
]

_FOLD = str.maketrans("ąćęłńóśźżĄĆĘŁŃÓŚŹŻ", "acelnoszzACELNOSZZ")


def fold(s: str) -> str:
    """ASCII-fold Polish diacritics and collapse whitespace."""
    return re.sub(r"\s+", " ", (s or "").translate(_FOLD)).strip()


def find_xlsx(session: requests.Session) -> tuple[str, str]:
    """Scrape the eRegistrations page -> (absolute xlsx url, period 'YYYY-MM')."""
    r = session.get(PAGE_URL, timeout=30)
    r.raise_for_status()
    m = re.search(r'href="([^"]*tabele[^"]*\.xlsx)"', r.text, re.IGNORECASE)
    if not m:
        sys.exit("Could not find the 'tabele ...xlsx' download link on the PZPM page.")
    href = m.group(1).replace("&amp;", "&")
    url = urllib.parse.urljoin(HOST, urllib.parse.quote(href, safe="/:?=&%"))
    pm = re.search(r"tabele\s*(\d{2})\.(\d{4})", urllib.parse.unquote(href))
    if not pm:
        sys.exit(f"Could not parse the period from the xlsx filename: {href}")
    return url, f"{pm.group(2)}-{pm.group(1)}"


def download_xlsx(session: requests.Session, url: str) -> bytes:
    r = session.get(url, headers={"Referer": PAGE_URL}, timeout=60)
    r.raise_for_status()
    return r.content


def parse_ogolem(xlsx_bytes: bytes) -> dict:
    """Parse the 'Ogółem' sheet -> {variant: {canonical col: value, 'TOTAL': t}}."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    if "Ogółem" not in wb.sheetnames:
        sys.exit(f"'Ogółem' sheet not found; sheets: {wb.sheetnames}")
    ws = wb["Ogółem"]

    header_to_variant = {fold(c["header"]).upper(): v for v, c in VARIANT_CONFIG.items()}
    result: dict[str, dict] = {}
    current = None  # variant currently being read, or None to skip

    for row in ws.iter_rows(values_only=True):
        label_raw = row[1] if len(row) > 1 else None       # col B
        if not isinstance(label_raw, str):
            continue
        label = fold(label_raw)
        upper = label.upper()

        if upper in header_to_variant:                     # category header
            current = header_to_variant[upper]
            total = row[2] if len(row) > 2 else None       # col C = current month
            result[current] = {"TOTAL": float(total) if isinstance(total, (int, float)) else 0.0}
            continue
        if upper in SKIP_HEADERS:                          # out-of-scope category
            current = None
            continue

        col = FUEL_MAP.get(label.lower())
        if current and col:
            val = row[2] if len(row) > 2 else None         # col C = current month
            if isinstance(val, (int, float)):
                result[current][col] = result[current].get(col, 0.0) + float(val)

    return result


def to_row(parsed_variant: dict, period: str, variant: str) -> dict | None:
    """Build one canonical CSV row from a parsed category dict."""
    total = parsed_variant.get("TOTAL", 0.0)
    if not total:
        return None
    core_cols = ["BEV", "PHEV", "HEV", "PETROL", "DIESEL"]
    core_sum = sum(parsed_variant.get(c, 0.0) for c in core_cols)
    row = {
        "period": period, "time_interval": "monthly",
        "variant": variant, "source": SOURCE,
        # "" = the category does not report this drive type separately.
        **{c: (parsed_variant[c] if c in parsed_variant else "") for c in core_cols},
        "OTHERS": max(0.0, total - core_sum),
        "TOTAL": total,
        "notes": "",
    }
    return row


def upsert_csv(csv_path: str, new_rows: dict) -> tuple[int, int]:
    existing: dict = {}
    if os.path.exists(csv_path):
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                for c in CSV_COLUMNS:
                    row.setdefault(c, "")
                existing[(row["period"], row["variant"])] = {k: row[k] for k in CSV_COLUMNS}

    added = updated = 0
    for key, new_row in sorted(new_rows.items()):
        if key not in existing:
            existing[key] = new_row
            added += 1
            print(f"  + {key[1]} {key[0]}")
        else:
            old = existing[key]
            for c in ["BEV", "PHEV", "HEV", "PETROL", "DIESEL", "OTHERS"]:
                ov = float(old.get(c) or 0)
                nv = float(new_row[c] or 0)
                if ov > 100 and abs(nv - ov) / ov > 0.5:
                    print(f"  WARNING {key[1]} {key[0]} {c}: existing={ov:.0f}, new={nv:.0f} "
                          f"— diff >50%, please verify")
            if not new_row.get("notes"):
                new_row["notes"] = old.get("notes", "")
            existing[key] = {**old, **new_row}
            updated += 1

    Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
    # CRLF to match the ACEA-family CSVs (data/Poland.csv, Belgium.csv, ...) so an
    # in-place update is a one-line diff rather than a whole-file re-write.
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLUMNS, lineterminator="\r\n")
        w.writeheader()
        for key in sorted(existing.keys(), key=lambda k: (k[1], k[0])):
            w.writerow(existing[key])
    return added, updated


def csv_has_period(csv_path: str, period: str, variant: str) -> bool:
    if not os.path.exists(csv_path):
        return False
    with open(csv_path, newline="", encoding="utf-8") as f:
        return any(r["period"] == period and r["variant"] == variant for r in csv.DictReader(f))


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--variant", choices=["whole", "vans", "hdv", "buses", "all"],
                    default="all", help="Which slice to fetch (default: all).")
    ap.add_argument("--xlsx", help="Parse a local workbook instead of scraping the page.")
    ap.add_argument("--period", help="Period 'YYYY-MM' (required with --xlsx).")
    ap.add_argument("--force", action="store_true",
                    help="Skip the 'period already present' early-exit.")
    args = ap.parse_args()

    aliases = {"whole": "Whole", "vans": "Vans", "hdv": "HDV", "buses": "Buses"}
    targets = list(aliases.values()) if args.variant == "all" else [aliases[args.variant]]

    if args.xlsx:
        if not args.period:
            ap.error("--period YYYY-MM is required with --xlsx")
        period = args.period
        xlsx_bytes = Path(args.xlsx).read_bytes()
    else:
        session = requests.Session()
        session.headers.update({"User-Agent": UA})
        url, period = find_xlsx(session)
        print(f"Latest PZPM workbook: {period}  ({url})")
        if not args.force:
            pending = [v for v in targets
                       if not csv_has_period(VARIANT_CONFIG[v]["csv"], period, v)]
            for v in [v for v in targets if v not in pending]:
                print(f"[{v}] CSV already has {period}; skipping.")
            targets = pending
            if not targets:
                print("All requested variants are current; nothing to do.")
                return
        xlsx_bytes = download_xlsx(session, url)

    parsed = parse_ogolem(xlsx_bytes)
    for variant in targets:
        row = to_row(parsed.get(variant, {}), period, variant)
        if row is None:
            print(f"[{variant}] no data for {period} in workbook; skipping.")
            continue
        cfg = VARIANT_CONFIG[variant]
        added, updated = upsert_csv(cfg["csv"], {(period, variant): row})
        print(f"[{variant}] {added} added, {updated} updated -> {cfg['csv']}")


if __name__ == "__main__":
    main()
