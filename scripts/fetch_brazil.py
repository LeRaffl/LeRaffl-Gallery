#!/usr/bin/env python3
"""
Fetch Brazil vehicle registration data from ANFAVEA and update data/Brazil.csv.

Usage
-----
    python scripts/fetch_brazil.py [--url URL] [--year YEAR] [--csv PATH]

* --url   Direct Excel URL or local file path. If omitted, the ANFAVEA
          index page is scraped for the current year's file.
* --year  Override the year (default: current calendar year).
* --csv   Target CSV (default: data/Brazil.csv).

This script is invoked by .github/workflows/fetch-brazil.yml on a monthly
cron (10th, 08:00 UTC) and via manual workflow_dispatch. When it produces
changes, the workflow commits data/Brazil.csv and triggers render-country.yml
for Brazil.

Data source
-----------
ANFAVEA (Associação Nacional dos Fabricantes de Veículos Automotores) publishes
one Excel workbook per year at https://anfavea.com.br/site/edicoes-em-excel/.
The canonical filename is `siteautoveiculos<YEAR>.xlsx`; mid-year revisions
may add a `-N` suffix. ANFAVEA also publishes `_nacionais` and `_importados`
splits — we explicitly ignore those (regex in `find_excel_url`) and parse
only the combined total file.

Parsing strategy
----------------
We read sheet "III. Emplacamento Combustível", which contains two tables:

  1. "Automóveis e Comerciais Leves" (cars + light commercial) — WE PARSE THIS
  2. "Caminhões e Ônibus" (trucks + buses) — IGNORED for now

Each table is laid out as:

    | <category> | <metric type> | Jan | Fev | Mar | ... | Dez | Total Ano |
    | Gasolina   |               | 6222| 5856| ...
    | Elétrico   |               | 8313| 8738| ...
    | ...

Row positions shift slightly year to year, so we locate the table dynamically:

  1. Find the FIRST row that contains the literal cell value "Unidades"
     (= the units sub-header of the cars table; "Porcenrtagem" [sic] marks
     the percentage table below it which we skip).
  2. The row immediately below that holds the month abbreviations
     (Jan, Fev, Mar, Abr, Mai, Jun, Jul, Ago, Set, Out, Nov, Dez). We build
     a {month_abbrev → column_index} map from whatever columns contain those
     strings, so layout changes don't break us.
  3. We then walk subsequent rows reading the fuel name from column B (or C
     as a fallback), looking it up in FUEL_MAP. Reading stops as soon as we
     hit a row containing "Fonte:" — that's the ANFAVEA end-of-table marker.

Column mapping (Portuguese → CSV column)
----------------------------------------
    Elétrico         → BEV
    Híbrido Plug-in  → PHEV
    Híbrido          → HEV         (regular non-plug-in hybrids)
    Gasolina         → PETROL
    Diesel           → DIESEL
    Flex Fuel        → FLEXFUEL    (Brazil-specific: gasoline + ethanol)
    (none)           → OTHERS = 0  (always zero for Brazil)
    sum of all above → TOTAL

Months without published data
-----------------------------
ANFAVEA pre-fills the entire calendar year with zeros and overwrites
month-by-month as data becomes available. We skip any month where ALL fuel
values are zero — that lets the script run mid-year without producing fake
"all-zero" rows for future months.

Upsert + plausibility check
---------------------------
Existing rows in data/Brazil.csv are keyed by `period` (YYYY-MM). For each
month parsed:

  * If the period is missing, we append a new row.
  * If it exists, we overwrite the row, but emit a WARNING to stdout if any
    fuel-type value moved by more than 50% versus the previous value — a
    cheap guard against parser drift or upstream relabeling.

The CSV is rewritten sorted by period. The `notes` field on touched rows
records the source filename (e.g. "siteautoveiculos2026.xlsx") for
provenance.

HTTP details
------------
ANFAVEA's Apache returns HTTP 406 Not Acceptable to the default
python-requests User-Agent, so we send desktop-Chrome headers (see
HTTP_HEADERS below) on both the page scrape and the xlsx download.
"""
import argparse
import csv
import io
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup

ANFAVEA_PAGE = "https://anfavea.com.br/site/edicoes-em-excel/"
TARGET_SHEET = "III. Emplacamento Combustível"

# ANFAVEA's Apache returns HTTP 406 for the default python-requests UA,
# so we identify as a regular desktop browser.
HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
}
CSV_COLUMNS = [
    "period", "time_interval", "variant", "source",
    "BEV", "PHEV", "HEV", "PETROL", "DIESEL", "FLEXFUEL",
    "OTHERS", "TOTAL", "notes",
]

# Maps Portuguese Excel labels → CSV column names
FUEL_MAP = {
    "Elétrico": "BEV",
    "Híbrido Plug-in": "PHEV",
    "Híbrido": "HEV",
    "Gasolina": "PETROL",
    "Diesel": "DIESEL",
    "Flex Fuel": "FLEXFUEL",
}

MONTH_ABR = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
             "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def find_excel_url(year: int) -> str:
    """Scrape the ANFAVEA page and return the xlsx URL for the given year.

    Prefers the canonical "siteautoveiculos<year>(-N)?.xlsx" (total file)
    over the "_nacionais" / "_importados" splits.
    """
    resp = requests.get(ANFAVEA_PAGE, headers=HTTP_HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # Matches siteautoveiculos2026.xlsx and siteautoveiculos2026-2.xlsx,
    # but NOT siteautoveiculos_nacionais2026.xlsx (digit must follow directly).
    pattern = re.compile(rf"siteautoveiculos{year}(?:[-_]\d+)?\.xlsx", re.IGNORECASE)
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if pattern.search(href):
            return href if href.startswith("http") else "https://anfavea.com.br" + href

    raise RuntimeError(
        f"Could not find ANFAVEA Excel URL for {year} on {ANFAVEA_PAGE}\n"
        "Pass --url to specify the file directly."
    )


def load_workbook_bytes(url_or_path: str) -> tuple[bytes, str]:
    """Return (file_bytes, source_filename). Accepts http(s) URL or local path."""
    if url_or_path.startswith("http://") or url_or_path.startswith("https://"):
        print(f"Downloading: {url_or_path}")
        resp = requests.get(url_or_path, headers=HTTP_HEADERS, timeout=60)
        resp.raise_for_status()
        return resp.content, url_or_path.split("/")[-1]
    path = url_or_path.replace("file://", "")
    with open(path, "rb") as f:
        return f.read(), Path(path).name


def parse_sheet(wb: openpyxl.Workbook, year: int) -> dict[str, dict]:
    """
    Parse the cars + light commercial "Unidades" table from the fuel sheet.
    Returns {period: row_dict} for months that have at least one non-zero value.
    """
    ws = wb[TARGET_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    # Find the first row where any cell equals "Unidades" — this is the
    # cars/light-commercial sub-header (trucks appear in a second table below)
    unidades_idx = None
    for i, row in enumerate(rows):
        if any(str(v).strip() == "Unidades" for v in row if v is not None):
            unidades_idx = i
            break
    if unidades_idx is None:
        raise RuntimeError(f"Could not find 'Unidades' header in sheet '{TARGET_SHEET}'")

    # The row immediately after contains month abbreviations as column headers
    month_row = rows[unidades_idx + 1]
    col_by_month: dict[str, int] = {
        str(v).strip(): i
        for i, v in enumerate(month_row)
        if v is not None and str(v).strip() in MONTH_ABR
    }
    if not col_by_month:
        raise RuntimeError("Could not locate month columns after 'Unidades' row")
    print(f"Month columns found: {list(col_by_month.keys())}")

    # Read fuel rows until "Fonte:" appears (end-of-table marker)
    fuel_data: dict[str, dict[str, float]] = {}
    for row in rows[unidades_idx + 2:]:
        # Fuel name sits in col B (index 1) or C (index 2)
        fuel_name = next(
            (str(v).strip() for v in row[1:3] if v is not None and str(v).strip() in FUEL_MAP),
            None,
        )
        if fuel_name is None:
            if any(v is not None and str(v).startswith("Fonte") for v in row):
                break
            continue
        fuel_data[fuel_name] = {
            month: float(row[col] if col < len(row) and row[col] is not None else 0)
            for month, col in col_by_month.items()
        }

    if not fuel_data:
        raise RuntimeError("No fuel rows found in 'Unidades' table")
    print(f"Fuel types found: {list(fuel_data.keys())}")

    # Build per-month CSV rows; skip months where all fuel values are zero
    result: dict[str, dict] = {}
    for month_idx, month_abr in enumerate(MONTH_ABR, start=1):
        if month_abr not in col_by_month:
            continue
        values = {fuel: fuel_data[fuel].get(month_abr, 0.0) for fuel in fuel_data}
        if all(v == 0.0 for v in values.values()):
            continue

        period = f"{year}-{month_idx:02d}"
        row: dict = {
            "period": period,
            "time_interval": "monthly",
            "variant": "Whole",
            "source": "ANFAVEA",
            "BEV": values.get("Elétrico", 0.0),
            "PHEV": values.get("Híbrido Plug-in", 0.0),
            "HEV": values.get("Híbrido", 0.0),
            "PETROL": values.get("Gasolina", 0.0),
            "DIESEL": values.get("Diesel", 0.0),
            "FLEXFUEL": values.get("Flex Fuel", 0.0),
            "OTHERS": 0.0,
        }
        row["TOTAL"] = sum(
            row[c] for c in ["BEV", "PHEV", "HEV", "PETROL", "DIESEL", "FLEXFUEL", "OTHERS"]
        )
        result[period] = row
    return result


def upsert_csv(csv_path: str, new_rows: dict[str, dict], source_filename: str) -> tuple[int, int]:
    """
    Upsert new_rows into csv_path by the 'period' key.
    Returns (added, updated) counts. Warns on implausible changes (>50% delta).
    """
    existing: dict[str, dict] = {}
    if os.path.exists(csv_path):
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                existing[row["period"]] = row

    added = updated = 0
    for period, new_row in sorted(new_rows.items()):
        new_row["notes"] = source_filename
        if period not in existing:
            existing[period] = new_row
            added += 1
            print(f"  + {period}")
        else:
            old = existing[period]
            for col in ["BEV", "PHEV", "HEV", "PETROL", "DIESEL", "FLEXFUEL"]:
                old_val = float(old.get(col) or 0)
                new_val = float(new_row[col])
                if old_val > 0 and abs(new_val - old_val) / old_val > 0.5:
                    print(
                        f"  WARNING {period} {col}: existing={old_val:.0f}, "
                        f"new={new_val:.0f} — diff >{50}%, please verify"
                    )
            existing[period] = {**old, **new_row}
            updated += 1
            print(f"  ~ {period}")

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, lineterminator="\n")
        writer.writeheader()
        for period in sorted(existing.keys()):
            writer.writerow(existing[period])

    return added, updated


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--url", help="Direct Excel URL or local file path")
    parser.add_argument("--year", type=int, default=datetime.now().year,
                        help="Year to fetch (default: current year)")
    parser.add_argument("--csv", default="data/Brazil.csv",
                        help="Path to Brazil CSV (default: data/Brazil.csv)")
    args = parser.parse_args()

    url = args.url
    if not url:
        print(f"Searching for {args.year} Excel on ANFAVEA page …")
        url = find_excel_url(args.year)
        print(f"Found: {url}")

    excel_bytes, source_filename = load_workbook_bytes(url)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    print(f"Opened workbook with sheets: {wb.sheetnames}")

    new_rows = parse_sheet(wb, args.year)
    if not new_rows:
        print("No new data found (all months are zero). Nothing to update.")
        sys.exit(0)
    print(f"Parsed {len(new_rows)} months: {sorted(new_rows.keys())}")

    added, updated = upsert_csv(args.csv, new_rows, source_filename)
    print(f"\nDone: {added} rows added, {updated} rows updated → {args.csv}")


if __name__ == "__main__":
    main()
